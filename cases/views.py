"""Views for the case workflow and the commercial master-data screens.

The workflow rules live in :mod:`cases.services`; these views only translate
HTTP requests into service calls, enforce who may see what, and render the
themed templates. Inbox visibility follows the business spec:

* Commercial sees the live status of *every* case (plus its reports).
* Technical / Supply only see the items currently sitting in their unit
  (their kartabl / inbox), sorted so pricing work is deadline-first.
"""
from __future__ import annotations

import json
import logging
from types import SimpleNamespace

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Prefetch, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.http import urlencode

from accounts.constants import Role, Unit

from . import codes, exports, services
from .constants import CaseStatus, FormKind, DocKind, OfferType, PriceType, EventAction, Side
from .forms import (CaseCreateForm, ClientForm, ClientRenameForm, CommentForm,
                    ExpertCodeForm)
from .inquiry_validate import validate_inquiry_rows
from .models import Case, CaseEvent, CaseForm, Client, ExpertCode, LineItem


def _vat_percent_value() -> float:
    try:
        from accounts.models import PlatformConfig
        return float(PlatformConfig.load().vat_percent or 10)
    except Exception:
        return 10.0

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def _profile(user):
    return getattr(user, "profile", None)


def _require_unit(user, *units) -> bool:
    p = _profile(user)
    return bool(p and not p.is_admin and p.unit in units)


def _parse_rows(form, files) -> list[dict]:
    """Return inquiry rows from the pasted grid JSON and/or an Excel upload.

    Both inputs map to the same four business columns: Item, Description,
    Size, Unit (Item is the row number). Excel parsing reads the first four
    columns regardless of their header text.

    IMPORTANT: the two inputs are mutually exclusive. The create form often
    populates the pasted-grid JSON FROM the uploaded Excel (for preview) and then
    submits BOTH, which previously appended the same rows twice (an N-row inquiry
    became 2N in the TO/PI). We therefore take the pasted grid when present and
    only fall back to the Excel file when there is no pasted grid.
    """
    rows: list[dict] = []

    pasted = (form.cleaned_data.get("pasted_table") or "").strip()
    if pasted:
        try:
            data = json.loads(pasted)
            for entry in data:
                if isinstance(entry, dict):
                    rows.append({
                        "client_row": entry.get("client_row") or entry.get("#") or "",
                        "description": entry.get("description") or entry.get("Description") or "",
                        "size": entry.get("size") or entry.get("Size") or "",
                        "unit": entry.get("unit") or entry.get("Unit") or "",
                        "quantity": entry.get("quantity") or entry.get("Qty") or "",
                    })
                elif isinstance(entry, (list, tuple)):
                    cells = list(entry) + ["", "", "", ""]
                    rows.append({
                        "description": cells[1], "size": cells[2],
                        "unit": cells[3], "quantity": "",
                    })
        except (ValueError, TypeError):
            pass

    # Only read the Excel upload when the pasted grid produced nothing, so the
    # same rows are never counted twice.
    if not rows:
        upload = files.get("excel_file")
        if upload:
            rows.extend(_rows_from_excel(upload))

    # Drop fully-empty rows.
    return [r for r in rows if any(
        str(v).strip() for k, v in r.items() if k != "client_row"
    )]


def _rows_from_excel(file_obj) -> list[dict]:
    import openpyxl

    rows: list[dict] = []
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    for idx, raw in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0 and _looks_like_header(raw):
            continue
        cells = [("" if c is None else str(c).strip()) for c in raw]
        # Drop a leading pure-integer row number if the file has one.
        if cells and cells[0].isdigit():
            cells = cells[1:]
        # Source columns are Description, Size, Qty, Unit (qty before unit).
        cells = (cells + ["", "", "", ""])[:4]
        rows.append({"description": cells[0], "size": cells[1],
                     "quantity": cells[2], "unit": cells[3]})
    wb.close()
    return rows


def _parse_jalali_deadline(raw: str):
    """Parse a Jalali 'YYYY-MM-DD[ HH:MM]' string into an aware datetime."""
    import datetime as _dt
    from django.utils import timezone
    from .jalali import jalali_to_gregorian
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        date_part, _, time_part = raw.partition(" ")
        # Accept dot, slash or dash as the date separator.
        norm = date_part.replace("/", "-").replace(".", "-")
        jy, jm, jd = [int(x) for x in norm.split("-")]
        if time_part:
            hh, mm = (int(x) for x in (time_part.split(":") + ["0", "0"])[:2])
        else:
            hh, mm = 0, 0
        gy, gm, gd = jalali_to_gregorian(jy, jm, jd)
        naive = _dt.datetime(gy, gm, gd, hh, mm)
    except (ValueError, TypeError):
        return None
    tz = timezone.get_current_timezone()
    return timezone.make_aware(naive, tz) if timezone.is_naive(naive) else naive



def _looks_like_header(raw) -> bool:
    """True when the first Excel row is a header (contains column titles like
    Description / Size / Qty / Unit) rather than real data, so we can skip it."""
    if not raw:
        return False
    text = " ".join(str(c).lower() for c in raw if c is not None)
    return any(k in text for k in ("description", "item", "size", "unit", "qty", "quantity"))


@login_required
def preview_excel(request):
    """Parse an uploaded Excel file and return its rows as JSON.

    Used by the new-case grid so the commercial user can see and edit the
    imported rows before creating the case.
    """
    from django.http import JsonResponse
    if request.method != "POST" or not request.FILES.get("excel_file"):
        return JsonResponse({"ok": False, "error": "No file uploaded."}, status=400)
    try:
        rows = _rows_from_excel(request.FILES["excel_file"])
    except Exception as exc:  # pragma: no cover - defensive
        return JsonResponse({"ok": False, "error": f"Could not read the file: {exc}"}, status=400)
    rows = [r for r in rows if any(r.values())]
    errs = validate_inquiry_rows(rows)
    if errs:
        return JsonResponse({
            "ok": False,
            "error": "Upload cancelled — inquiry rules failed:\n" + "\n".join(errs),
            "errors": errs,
        }, status=400)
    return JsonResponse({"ok": True, "rows": rows})


# ---------------------------------------------------------------------------
# Inbox (kartabl)
# ---------------------------------------------------------------------------
@login_required
def inbox(request):
    profile = _profile(request.user)
    if profile is None or profile.is_admin:
        return redirect("accounts:admin_console")
    if profile.is_general_manager:
        return redirect("reports:dashboard")

    from people.role_nav import work_context
    ctx = work_context(request)
    role = ctx.role
    unit = (role.unit if role is not None else profile.unit) or profile.unit
    # Membership is computed by the shared helper (split-aware) so the inbox and
    # the nav badge always agree; here we only add the per-unit ordering.
    cases = services.inbox_cases_for_request(request).select_related("client")

    # Inbox ordering for EVERY unit: the nearest deadline sits at the top; cases
    # without a deadline sink to the bottom, ordered by creation time.
    cases = cases.extra(select={"no_deadline": "deadline IS NULL"}).order_by(
        "no_deadline", "deadline", "created_at")
    scope = {Unit.SUPPLY: "supply", Unit.TECHNICAL: "technical",
             Unit.COMMERCIAL: "commercial"}.get(unit, "none")
    context_extra = {"scope": scope, "unit": unit}

    # Status summary over the whole inbox (grouped COUNT — not by loading rows).
    from django.db.models import Count
    summary = {}
    for row in (services.inbox_cases_for_request(request)
                .order_by().values("status").annotate(n=Count("id", distinct=True))):
        label = CaseStatus.LABELS.get(row["status"], row["status"])
        summary[label] = summary.get(label, 0) + row["n"]

    # No pagination: every inbox case is rendered once; the table scrolls via
    # .vscroll (same sticky-header pattern as inquiry / TO / PI tables).
    cases_list = list(cases)
    for case in cases_list:
        case.inbox_status_rows = services.inbox_status_rows(case, ctx.seat_user)

    fx_stale = False
    is_manager = (
        (role.role if role is not None else profile.role) == Role.MANAGER
    )
    if is_manager and unit == Unit.COMMERCIAL:
        try:
            from .fx_rates import is_rates_stale
            fx_stale = is_rates_stale()
        except Exception:
            fx_stale = False

    from people.role_nav import role_can_create_case
    can_create = False
    if role is not None:
        can_create = role_can_create_case(role, is_substitute=ctx.is_substitute)
    elif not ctx.is_substitute:
        can_create = profile.can_create_case

    context = {
        "cases": cases_list,
        "summary": {k: v for k, v in summary.items() if v},
        "can_create": can_create,
        "fx_stale": fx_stale,
        "is_substitute": ctx.is_substitute,
        **context_extra,
    }
    return render(request, "cases/inbox.html", context)


# ---------------------------------------------------------------------------
# Case creation
# ---------------------------------------------------------------------------
@login_required
def archive(request):
    """Searchable history of cases.

    Commercial users (and admins) may search every case; Technical/Supply see
    the cases that have passed through their hands. Search matches the global
    serial number, the client name or the client code, and can be filtered by
    status and by creation date.

    Archive always includes every case that is also in the user's inbox
    (and drafts), so active files are never missing from history.
    """
    profile = _profile(request.user)
    if profile is None:
        return redirect("accounts:login")

    select_mode = str(request.GET.get("select") or "").strip() in ("1", "true", "yes")
    select_return = (request.GET.get("return") or "").strip()
    if select_mode and not select_return.startswith("/"):
        select_return = ""

    qs = Case.objects.select_related("client", "created_by").all()

    # Manager "Only my cases" toggle (default OFF = current wide archive view).
    mine_only = str(request.GET.get("mine", "") or "").strip() in ("1", "true", "on", "yes")
    is_unit_manager = bool(
        profile.role == Role.MANAGER
        and profile.unit in (Unit.COMMERCIAL, Unit.TECHNICAL, Unit.SUPPLY)
        and not profile.is_admin
        and not profile.is_general_manager
    )

    # Scope:
    #   admin / general manager     -> the entire archive
    #   commercial manager (off)    -> the entire archive (incl. experts' cases)
    #   commercial manager (on)     -> only cases they personally created
    #   commercial expert           -> only cases they personally created
    #   technical manager (on)      -> only TOs they personally created
    #   supply manager (on)         -> only PIs they personally created
    #   unit supervisor             -> every case that passed through their unit
    #   technical / supply (else)   -> cases they personally participated in
    # Scope uses the active seat user so secondary / translated seats see
    # their own cases (not the login profile's primary seat alone).
    from people.role_nav import work_context
    ctx = work_context(request)
    seat_user = ctx.seat_user
    if ctx.role is not None:
        # Prefer active role for archive unit/role filters.
        profile_unit = ctx.role.unit or profile.unit
        profile_role = ctx.role.role or profile.role
    else:
        profile_unit = profile.unit
        profile_role = profile.role

    if profile.is_admin or profile.is_general_manager:
        pass
    elif profile_unit == Unit.COMMERCIAL and profile_role == Role.MANAGER:
        if mine_only:
            qs = qs.filter(created_by=seat_user)
    elif profile_unit == Unit.COMMERCIAL:
        qs = qs.filter(created_by=seat_user)
        # Substitutes do not see fully closed / terminal archive rows.
        if ctx.is_substitute:
            qs = qs.exclude(status__in=CaseStatus.ENDED)
    elif is_unit_manager and mine_only and profile_unit == Unit.TECHNICAL:
        qs = qs.filter(
            forms__kind=FormKind.TO, forms__created_by=seat_user
        ).distinct()
    elif is_unit_manager and mine_only and profile_unit == Unit.SUPPLY:
        qs = qs.filter(
            forms__kind=FormKind.PI, forms__created_by=seat_user
        ).distinct()
    elif profile_role == Role.SUPERVISOR and profile_unit:
        u = profile_unit
        qs = qs.filter(
            Q(holder_unit=u)
            | Q(events__from_unit=u) | Q(events__to_unit=u)
            | Q(forms__unit_at_creation=u)
        ).distinct()
    else:
        qs = qs.filter(
            Q(created_by=seat_user)
            | Q(assigned_to=seat_user)
            | Q(forms__created_by=seat_user)
            | Q(events__actor=seat_user)
        ).distinct()
        if ctx.is_substitute:
            qs = qs.exclude(status__in=CaseStatus.ENDED)

    # Always include this user's current inbox (and therefore drafts / live
    # files sitting there) so Archive ⊇ Inbox — except when the manager
    # "Only my cases" toggle is ON (that view is intentionally narrower).
    if not mine_only:
        try:
            inbox_qs = services.inbox_cases_for_request(request)
            if inbox_qs is not None:
                qs = (qs | inbox_qs).distinct()
        except Exception:
            logger.exception("archive: failed to union inbox cases for user %s", request.user.pk)

    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    doc_kind = request.GET.get("kind", "").strip()
    date_from = request.GET.get("from", "").strip()
    date_to = request.GET.get("to", "").strip()

    # Drill-down from the dashboard: a single expert's cases.
    creator_id = request.GET.get("creator", "").strip()
    assignee_id = request.GET.get("assignee", "").strip()
    drill_person = None
    if creator_id.isdigit():
        qs = qs.filter(created_by_id=int(creator_id))
        drill_person = User.objects.filter(pk=int(creator_id)).first()
    elif assignee_id.isdigit():
        aid = int(assignee_id)
        qs = qs.filter(
            Q(technical_assignee_id=aid) | Q(supply_internal_assignee_id=aid)
            | Q(supply_external_assignee_id=aid) | Q(supply_assignee_id=aid)
            | Q(assigned_to_id=aid)
        ).distinct()
        drill_person = User.objects.filter(pk=aid).first()

    if query:
        flt = Q(client__name__icontains=query) | Q(client__code__icontains=query) | Q(doc_no__icontains=query)
        if query.isdigit():
            flt |= Q(serial=int(query))
        qs = qs.filter(flt)
    if status:
        codes = [c for c in status.split(",") if c]
        qs = qs.filter(status__in=codes) if codes else qs
    if doc_kind:
        kinds = [k for k in doc_kind.split(",") if k]
        qs = qs.filter(kind__in=kinds) if kinds else qs
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    qs = qs.order_by("-created_at")

    # Collapsed status groups (so "With Technical" covers both WITH_TECHNICAL and
    # RETURNED_TO_TECHNICAL, etc.) used by status tabs + row data-fval.
    status_group = CaseStatus.ARCHIVE_GROUP

    # Build the filter dropdown choices from the WHOLE filtered set using cheap
    # DISTINCT queries — not by loading every case — so the options stay complete
    # even though only one page of rows is rendered below.
    def _offer_label(offer_type, upgraded):
        if offer_type != OfferType.TO_PI:
            return "TO"
        return "TO & PI (Two Stage)" if upgraded else "TO & PI"

    f_clients = sorted({
        f"{name} ({code})"
        for name, code in qs.values_list("client__name", "client__code").distinct()
        if name
    })
    f_experts = sorted({
        f"{((first + ' ' + last).strip() or username)} ({ecode})"
        for first, last, username, ecode in qs.values_list(
            "created_by__first_name", "created_by__last_name",
            "created_by__username", "expert_code").distinct()
    })
    f_prices = sorted({
        PriceType.LABELS.get(pt, "")
        for pt in qs.values_list("price_type", flat=True).distinct() if pt
    })
    f_offers = sorted({
        _offer_label(ot, up)
        for ot, up in qs.values_list("offer_type", "upgraded_two_stage").distinct()
    })
    f_kinds = sorted({
        dict(DocKind.CHOICES).get(k, k)
        for k in qs.values_list("kind", flat=True).distinct() if k
    })
    f_orders = sorted({
        ono for ono in qs.values_list("order_no", flat=True).distinct()
        if (ono or "").strip()
    })

    # No pagination: every matching case is rendered once; the table scrolls
    # via .vscroll (sticky header), same idea as inquiry / TO / PI tables.
    cases_list = list(qs)
    tab_counts = {label: 0 for label in CaseStatus.ARCHIVE_TAB_ORDER}
    for c in cases_list:
        if c.is_split:
            groups = {status_group.get(c.side_status(sc), c.status_label) for sc in c.sides}
        else:
            groups = {status_group.get(c.status, c.status_label)}
        c.status_fval = " ".join(sorted(groups))
        for g in groups:
            if g in tab_counts:
                tab_counts[g] += 1

    status_tabs = [
        {
            "label": label,
            "count": tab_counts.get(label, 0),
            "color": CaseStatus.ARCHIVE_TAB_COLORS.get(label, "#64748b"),
            "words": label.split(),
        }
        for label in CaseStatus.ARCHIVE_TAB_ORDER
        if tab_counts.get(label, 0) > 0
    ]

    # PI grand totals (VAT-inclusive) for every visible row + drill-down sum.
    show_money = bool(
        profile.is_admin or profile.is_general_manager
        or profile.unit == Unit.COMMERCIAL
    )
    drill_grand_total_display = ""
    if show_money:
        from .export_data import case_pi_grand_totals_map, format_money_amount
        all_ids = [c.pk for c in cases_list]
        gt_map = case_pi_grand_totals_map(all_ids)
        for c in cases_list:
            amt = gt_map.get(c.pk, 0.0)
            c.grand_total_num = amt
            c.grand_total_display = format_money_amount(amt) if amt else "—"
        if drill_person:
            drill_sum = sum(gt_map.values()) if gt_map else 0.0
            drill_grand_total_display = (
                format_money_amount(drill_sum) if drill_sum else "—"
            )

    show_expert_filter = bool(
        profile and (profile.role in (Role.SUPERVISOR, Role.MANAGER)
                     or profile.is_general_manager or profile.is_admin))
    mine_toggle_label = {
        Unit.COMMERCIAL: "Only cases I created",
        Unit.TECHNICAL: "Only TOs I created",
        Unit.SUPPLY: "Only PIs I created",
    }.get(profile.unit or "", "Only my cases")
    return render(request, "cases/archive.html", {
        "cases": cases_list,
        "total_count": len(cases_list),
        "query": query,
        "status": status,
        "doc_kind": doc_kind,
        "date_from": date_from,
        "date_to": date_to,
        "status_choices": CaseStatus.CHOICES,
        "kind_choices": DocKind.CHOICES,
        "unit": "" if (profile.is_admin or profile.is_general_manager) else profile.unit,
        "is_admin": profile.is_admin or profile.is_general_manager,
        "show_expert_filter": show_expert_filter,
        "viewer_unit": profile.unit,
        "drill_person": drill_person,
        "show_money": show_money,
        "drill_grand_total_display": drill_grand_total_display,
        "f_clients": f_clients,
        "f_experts": f_experts,
        "f_prices": f_prices,
        "f_offers": f_offers,
        "f_kinds": f_kinds,
        "f_orders": f_orders,
        "status_tabs": status_tabs,
        "show_mine_toggle": is_unit_manager,
        "mine_only": mine_only,
        "mine_toggle_label": mine_toggle_label,
        "select_mode": select_mode,
        "select_return": select_return,
    })


# ---------------------------------------------------------------------------
# Case creation
# ---------------------------------------------------------------------------
@login_required
def case_create(request):
    profile = _profile(request.user)
    from people.role_nav import role_can_create_case, work_context
    ctx = work_context(request)
    can_create = False
    if ctx.role is not None:
        can_create = role_can_create_case(ctx.role, is_substitute=ctx.is_substitute)
    elif profile is not None and not ctx.is_substitute:
        can_create = profile.can_create_case
    if profile is None or not can_create:
        if ctx.is_substitute:
            messages.error(request, "Substitutes cannot open a new case. Return the seat first.")
        else:
            messages.error(request, "Only Commercial users can open a new case.")
        return redirect("cases:inbox")

    if request.method == "POST":
        form = CaseCreateForm(request.POST, request.FILES)
        if form.is_valid():
            rows = _parse_rows(form, request.FILES)
            if not rows:
                form.add_error(None, "Add at least one item to the inquiry table.")
            else:
                inq_errs = validate_inquiry_rows(rows)
                if inq_errs:
                    for err in inq_errs:
                        form.add_error(None, err)
                else:
                    try:
                        case = services.create_case(
                            creator=ctx.seat_user,
                            kind=form.cleaned_data["kind"],
                            offer_type=form.cleaned_data["offer_type"],
                            client=form.cleaned_data["client"],
                            order_no=form.cleaned_data.get("order_no", ""),
                            deadline=form.cleaned_data.get("deadline"),
                            price_type=form.cleaned_data.get("price_type", "INTERNAL"),
                            client_commercial_expert=form.cleaned_data.get("client_commercial_expert", ""),
                            client_commercial_phone=form.cleaned_data.get("client_commercial_phone", ""),
                            client_technical_expert=form.cleaned_data.get("client_technical_expert", ""),
                            client_technical_phone=form.cleaned_data.get("client_technical_phone", ""),
                            rows=rows,
                        )
                    except ValueError as exc:
                        form.add_error(None, str(exc))
                    else:
                        messages.success(request, f"Case {case.doc_no} created.")
                        return redirect("cases:case_detail", pk=case.pk)
    else:
        form = CaseCreateForm()

    return render(request, "cases/case_create.html", {"form": form})


# ---------------------------------------------------------------------------
# Case detail
# ---------------------------------------------------------------------------
def _fmt_duration(seconds) -> str:
    """Human duration that never shows '0 days'.

    >= 1 day  -> 'Xd Yh'   (hours dropped when zero)
    >= 1 hour -> 'Xh Ym'   (minutes dropped when zero)
    else      -> 'Xm'
    """
    seconds = int(max(0, seconds or 0))
    days, rem = divmod(seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes = rem // 60
    if days:
        return f"{days}d {hours}h" if hours else f"{days}d"
    if hours:
        return f"{hours}h {minutes}m" if minutes else f"{hours}h"
    return f"{minutes}m"


def _case_lifecycle_report(case):
    """Lifecycle report(s) for a case (managers/admin only).

    Non-combined  -> a single card (Commercial / Technical / Supply).
    Combined       -> two cards, one per side (Internal / External), each a full
                      independent report for that side with its own people,
                      per-unit durations, item count and total.

    Timing rules:
    * Unit clocks (Commercial / Technical / Supply) accumulate while that unit
      holds the case. They pause while status is CLOSED (With Client) or
      FINAL_APPROVED — those phases have their own rows.
    * With Client runs only while status is CLOSED (from CLOSE until reopen /
      finalize / burn / final-close). A NEW_VERSION after sent-to-client stops
      the client clock; it must not keep ticking after the summary changes.
    * Final Approved runs only while status is FINAL_APPROVED.
    """
    from collections import defaultdict
    from django.utils import timezone
    now = timezone.now()
    events = list(case.events.order_by("created_at", "id"))

    # Statuses whose time is reported on dedicated rows, not on unit clocks.
    UNIT_CLOCK_PAUSE = {CaseStatus.CLOSED, CaseStatus.FINAL_APPROVED}

    # Sent-to-client ends when the case is reopened (new version), finalized,
    # burned, final-closed, or cancelled — not only on finalize/burn.
    CLIENT_END = {
        EventAction.NEW_VERSION,
        EventAction.FINALIZE,
        EventAction.BURN,
        EventAction.FINAL_CLOSE,
        EventAction.CANCEL,
        EventAction.APPROVE_CANCEL,
    }
    FINAL_END = {
        EventAction.FINAL_CLOSE,
        EventAction.BURN,
        EventAction.NEW_VERSION,
        EventAction.CANCEL,
        EventAction.APPROVE_CANCEL,
    }

    def _status_of(side):
        if side is None:
            return case.status
        return case.side_status(side)

    def unit_seconds(side=None):
        evs = events if side is None else [e for e in events if (e.side or "") in (side, "")]
        secs = defaultdict(float)
        # "client" / "final" phases are timed on their own rows — skip them here.
        phase = None
        for i in range(len(evs) - 1):
            e = evs[i]
            if e.action == EventAction.CLOSE:
                phase = "client"
            elif e.action == EventAction.FINALIZE:
                phase = "final"
            elif phase == "client" and e.action in CLIENT_END:
                phase = None
            elif phase == "final" and e.action in FINAL_END:
                phase = None

            if phase is not None:
                continue
            holder = e.to_unit or e.from_unit
            if holder:
                secs[holder] += (evs[i + 1].created_at - e.created_at).total_seconds()

        if evs:
            cur_status = _status_of(side)
            is_terminal = cur_status in CaseStatus.TERMINAL
            # Live unit clock only while actively with a unit (yellow/blue/etc.).
            if (not is_terminal) and cur_status not in UNIT_CLOCK_PAUSE:
                if side is None:
                    cur_holder = case.holder_unit
                else:
                    cur_holder = case.side_holder(side)
                holder = evs[-1].to_unit or evs[-1].from_unit or cur_holder
                if holder:
                    secs[holder] += (now - evs[-1].created_at).total_seconds()
        return secs

    def author(kind, side=None):
        qs = case.forms.filter(kind=kind)
        if side is not None:
            qs = qs.filter(side=side)
        f = qs.order_by("version", "id").first()
        if f and f.created_by:
            return f.created_by.get_full_name() or f.created_by.username
        return "—"

    creator = (case.created_by.get_full_name() or case.created_by.username) if case.created_by else "—"
    item_count = case.line_items.count()

    def phase_seconds(side, start_actions, end_actions, *, live_status):
        """Sum every start→end interval; keep ticking only while live_status."""
        evs = events if side is None else [e for e in events if (e.side or "") in (side, "")]
        total = 0.0
        start_at = None
        for e in evs:
            if start_at is None and e.action in start_actions:
                start_at = e.created_at
            elif start_at is not None and e.action in end_actions:
                total += (e.created_at - start_at).total_seconds()
                start_at = None
        if start_at is not None and _status_of(side) == live_status:
            total += (now - start_at).total_seconds()
        return total

    def make_card(side, label):
        secs = unit_seconds(side)
        cur_status = _status_of(side)
        is_terminal = cur_status in CaseStatus.TERMINAL
        if side is None:
            end_time = events[-1].created_at if (is_terminal and events) else now
        else:
            side_evs = [e for e in events if (e.side or "") in (side, "")]
            end_time = side_evs[-1].created_at if (is_terminal and side_evs) else now
        total_seconds = (end_time - case.created_at).total_seconds() if events else 0

        with_client = phase_seconds(
            side, {EventAction.CLOSE}, CLIENT_END, live_status=CaseStatus.CLOSED)
        final_appr = phase_seconds(
            side, {EventAction.FINALIZE}, FINAL_END, live_status=CaseStatus.FINAL_APPROVED)

        rows = [
            {"unit": "COMMERCIAL", "label": "Commercial", "person": creator,
             "duration": _fmt_duration(secs.get(Unit.COMMERCIAL, 0))},
            {"unit": "TECHNICAL", "label": "Technical", "person": author(FormKind.TO, side),
             "duration": _fmt_duration(secs.get(Unit.TECHNICAL, 0))},
            {"unit": "SUPPLY", "label": "Supply", "person": author(FormKind.PI, side),
             "duration": _fmt_duration(secs.get(Unit.SUPPLY, 0))},
            {"unit": "CLIENT", "label": "With Client", "person": "—",
             "duration": _fmt_duration(with_client)},
            {"unit": "FINAL", "label": "Final Approved", "person": "—",
             "duration": _fmt_duration(final_appr)},
        ]
        return {"side_label": label, "item_count": item_count, "created_at": case.created_at,
                "rows": rows, "total": _fmt_duration(total_seconds), "is_terminal": is_terminal}

    if case.is_split:
        cards = [make_card(Side.INTERNAL, "Internal"), make_card(Side.EXTERNAL, "External")]
    else:
        cards = [make_card(None, "")]
    return {"cards": cards}


def _side_notes(events_recent, side_code, holder_unit):
    """Return (arrival_comment, assign_comment) for one side of a split case.

    Same idea as the case-level notes, but restricted to events tagged with this
    side (legacy untagged events count for both).
    """
    side_events = [e for e in events_recent if (e.side or "") in (side_code, "")]
    arrival = None
    handoff = None
    for e in side_events:
        if e.to_unit == holder_unit and e.from_unit and e.from_unit != e.to_unit:
            handoff = e
            break
    if handoff is not None:
        if handoff.comment:
            arrival = handoff
        else:
            sending = handoff.from_unit
            for e in side_events:
                if (e.comment and e.from_unit == sending
                        and e.created_at <= handoff.created_at
                        and e.action != EventAction.ASSIGN):
                    arrival = e
                    break
    assign = None
    for e in side_events:
        if e.action == EventAction.ASSIGN and e.comment and e.to_unit == holder_unit:
            assign = e
            break
    return arrival, assign


@login_required
def case_detail(request, pk):
    case = (Case.objects
            .select_related("client", "created_by", "assigned_to")
            # Pull every form and event once; current_form()/services read these
            # from the prefetch cache, turning dozens of per-form queries into two.
            .prefetch_related(
                Prefetch("forms", queryset=CaseForm.objects.select_related("created_by")),
                Prefetch("events",
                         queryset=CaseEvent.objects.select_related("actor", "actor__profile")))
            .filter(pk=pk)
            .first())
    if case is None:
        # The case isn't in the database (e.g. a stale/bookmarked URL pointing at
        # a row that no longer exists). Send the user to their inbox with a short
        # note instead of showing a raw 404 page.
        messages.info(request, f"Case #{pk} no longer exists.")
        return redirect("cases:inbox")
    profile = _profile(request.user)
    from people.role_nav import work_context
    ctx = work_context(request)
    role = ctx.role
    seat_user = ctx.seat_user
    seat_id = getattr(seat_user, "id", None)
    # Active PersonRole drives unit/role for secondary seats (login profile stays
    # on the primary seat and must NOT scope forms / timeline / action chrome).
    if role is not None:
        viewer_unit = (role.unit or "").strip() or (profile.unit if profile else None)
        viewer_role = (role.role or "").strip() or (profile.role if profile else None)
        is_admin_view = bool(
            role.is_admin or role.is_general_manager
            or (profile and (profile.is_admin or profile.is_general_manager))
        )
    else:
        viewer_unit = profile.unit if profile else None
        viewer_role = profile.role if profile else None
        is_admin_view = bool(profile and (profile.is_admin or profile.is_general_manager))
    is_supply = viewer_unit == Unit.SUPPLY
    is_tech = viewer_unit == Unit.TECHNICAL
    is_comm = viewer_unit == Unit.COMMERCIAL
    is_mgr = viewer_role == Role.MANAGER

    actions = (
        services.allowed_actions(
            case, request.user, role=role, work_user=seat_user,
        ) if profile else set()
    )

    # In-memory views of the prefetched forms/events (no extra queries). All the
    # repeated filtering below runs over these lists instead of hitting the DB.
    case_forms = list(case.forms.all())
    case_events = list(case.events.all())

    def _forms_of(kind, side=None, sent=None):
        out = [f for f in case_forms if f.kind == kind]
        if side is not None:
            out = [f for f in out if f.side == side]
        if sent is not None:
            out = [f for f in out if bool(f.sent) == sent]
        return out

    # Admins may always inspect a case read-only. Everyone else may view a case
    # they participated in (created, were assigned, authored a form, or acted on),
    # plus Commercial who can view any case, plus a unit MANAGER who can always
    # see cases that belong to (or passed through) their own unit. The actual
    # rule lives in services.user_can_view_case (shared with the export routes
    # below, which used to skip this check entirely).
    if not services.user_can_view_case(case, request.user,
                                        case_forms=case_forms, case_events=case_events,
                                        role=role, work_user=seat_user):
        messages.error(request, "You do not have access to this case.")
        return redirect("cases:inbox")

    def _hide_fx_only(forms):
        """Technical/Supply never see Commercial-only currency-conversion snapshots.

        Those versions exist for Commercial Proforma FX only and must not appear
        as TO/PI/Inquiry chips for Technical or Supply (they keep their prior
        real versions; the next real revision is e.g. 04, skipping FX-only 03).
        """
        if is_admin_view or is_comm:
            return forms
        return [f for f in forms if not services.form_is_currency_conversion_only(f)]

    def _display_current(kind, side=None):
        """Current form for UI; Technical/Supply skip FX-only Commercial clones."""
        cur = case.current_form(kind, side)
        if cur and services.form_is_currency_conversion_only(cur):
            if not (is_admin_view or is_comm):
                reals = _hide_fx_only(_forms_of(kind, side=side))
                cur = max(reals, key=lambda f: (f.version, f.id), default=None) if reals else None
        return cur

    forms_by_kind = {
        FormKind.INQUIRY: _display_current(FormKind.INQUIRY),
        FormKind.TO: _display_current(FormKind.TO),
        FormKind.PI: _display_current(FormKind.PI),
    }
    # Each unit sees every version of the form it owns and only the latest
    # version published *to them* for the others; admins/GMs see every version.
    owner_unit = {FormKind.INQUIRY: Unit.COMMERCIAL,
                  FormKind.TO: Unit.TECHNICAL,
                  FormKind.PI: Unit.SUPPLY}

    def _form_chip_sort_key(f):
        """Left → right: older versions first; same version by creation time.

        Keeps ``Version 00`` left of ``Version 00 · Two Stage`` (and any later
        same-number snapshot) so each newly built form lands to the right.
        """
        created = getattr(f, "created_at", None)
        # Naive fallback keeps None sorted first among equals.
        return (int(getattr(f, "version", 0) or 0), created is not None, created, int(getattr(f, "pk", 0) or 0))

    def _forms_published_to_viewer(kind, side=None):
        """Non-owner: latest form version handed to this viewer's unit."""
        forms = _hide_fx_only(_forms_of(kind, side=side))
        visible = [
            f for f in forms
            if services.form_published_to_unit(f, viewer_unit)
        ]
        visible = sorted(visible, key=_form_chip_sort_key)
        return visible[-1:] if visible else []

    version_lists = {}
    for kind in (FormKind.INQUIRY, FormKind.TO, FormKind.PI):
        if is_admin_view or viewer_unit == owner_unit[kind]:
            # Owning unit (incl. its manager) and admin/GM see every version
            # — except Technical/Supply never see FX-only Commercial clones.
            version_lists[kind] = sorted(
                _hide_fx_only(_forms_of(kind)), key=_form_chip_sort_key)
        else:
            version_lists[kind] = _forms_published_to_viewer(kind)
    versions = {
        FormKind.TO: sorted(_hide_fx_only(_forms_of(FormKind.TO)),
                            key=_form_chip_sort_key, reverse=True),
        FormKind.PI: sorted(_hide_fx_only(_forms_of(FormKind.PI)),
                            key=_form_chip_sort_key, reverse=True),
    }

    # Possible assignees when the manager wants to delegate the case.
    assignees = []
    if "assign" in actions:
        assignees = User.objects.filter(
            profile__unit=case.holder_unit,
            profile__role=Role.EXPERT,
            is_active=True,
        ).select_related("profile").order_by("first_name", "username")

    # Timeline scoping: a unit sees its own internal actions plus the handoff
    # events that touch it (received/sent). Admins see the full timeline.
    all_events = case_events
    if is_admin_view:
        events = list(all_events)
    else:
        events = [e for e in all_events if _event_visible_to(e, viewer_unit)]

    # The handoff note shown in the action panel: the last thing the unit that
    # sent the case here said. Either the comment attached to the send action,
    # or — if they commented separately and then sent — that unit's most recent
    # comment made at/just before the handoff. Shown in the sending unit's colour.
    events_recent = sorted(
        all_events, key=lambda e: (e.created_at, e.id), reverse=True
    )
    arrival_comment = None
    handoff = None
    for e in events_recent:
        if e.to_unit == case.holder_unit and e.from_unit and e.from_unit != e.to_unit:
            handoff = e
            break
    if handoff is not None:
        if handoff.comment:
            arrival_comment = handoff
        else:
            sending_unit = handoff.from_unit
            for e in events_recent:
                if (e.comment and e.from_unit == sending_unit
                        and e.created_at <= handoff.created_at
                        and e.action != EventAction.ASSIGN):
                    arrival_comment = e
                    break

    # The latest assignment note made WITHIN the current holder unit (so a
    # previous unit's manager note never leaks to the next unit).
    assign_comment = None
    for e in events_recent:
        if (e.action == EventAction.ASSIGN and e.comment
                and e.to_unit == case.holder_unit):
            assign_comment = e
            break

    # The technical expert who first built the TO (shown in case information).
    to_author = None
    first_to = min(
        (f for f in _forms_of(FormKind.TO) if not services.form_is_currency_conversion_only(f)),
        key=lambda f: (f.version, f.id), default=None)
    if first_to:
        to_author = first_to.created_by

    # Tool mode for the Build buttons: build a first form, edit the current
    # (still-unsent) form, or branch a new version once it has been sent.
    def _form_mode(kind, side=None):
        cur = case.current_form(kind, side)
        # Commercial FX-only clones are not Technical/Supply work — treat the
        # latest real snapshot as current for mode decisions.
        if cur and services.form_is_currency_conversion_only(cur):
            reals = [
                f for f in _forms_of(kind, side=side)
                if not services.form_is_currency_conversion_only(f)
            ]
            cur = max(reals, key=lambda f: (f.version, f.id), default=None) if reals else None
        if not cur:
            return "build"
        inq = case.current_form(FormKind.INQUIRY, side)
        # A new inquiry version (beyond this form's version) — OR a new two-stage
        # generation at the SAME number (the case was upgraded to TO & PI two
        # stage) — forces a matching new form version. Otherwise the unit edits
        # its current form in place — even after sending it and getting the case
        # back, as long as the inquiry has not moved on.
        if inq and (cur.version < inq.version
                    or bool(cur.two_stage) != bool(inq.two_stage)):
            return "newversion"
        return "edit"
    to_mode = _form_mode(FormKind.TO)
    pi_mode = _form_mode(FormKind.PI)

    # ---- Per-side bundles (Internal / External sub-streams) -----------
    side_codes = case.sides or [""]
    supply_assignee_of = {
        Side.INTERNAL: case.supply_internal_assignee_id,
        Side.EXTERNAL: case.supply_external_assignee_id,
        "": case.supply_assignee_id,
    }
    is_supply = viewer_unit == Unit.SUPPLY
    is_mgr = viewer_role == Role.MANAGER
    is_tech = viewer_unit == Unit.TECHNICAL
    is_comm = viewer_unit == Unit.COMMERCIAL
    tech_owns = bool(is_mgr or case.technical_assignee_id == seat_id)
    # Editing items re-snapshots every side's inquiry, so it must stop once any
    # side's inquiry has been sent (otherwise a sent side would be overwritten).
    inq_any_sent = bool(_forms_of(FormKind.INQUIRY, sent=True))
    sides_data = []
    for sc in side_codes:
        s_forms = {
            FormKind.INQUIRY: _display_current(FormKind.INQUIRY, sc),
            FormKind.TO: _display_current(FormKind.TO, sc),
            FormKind.PI: _display_current(FormKind.PI, sc),
        }
        s_vlists = {}
        for kind in (FormKind.INQUIRY, FormKind.TO, FormKind.PI):
            if is_admin_view or viewer_unit == owner_unit[kind]:
                # Owning unit (incl. its manager) and admin/GM see every version
                # — except Technical/Supply never see FX-only Commercial clones.
                s_vlists[kind] = sorted(
                    _hide_fx_only(_forms_of(kind, side=sc)), key=_form_chip_sort_key)
            else:
                s_vlists[kind] = _forms_published_to_viewer(kind, side=sc)
        # Per-side supply permission. For split cases each side is independent:
        # a user acts on a side only while that side is still at Supply and they
        # own it (or, for an un-delegated side, they are the supply manager).
        s_assignee = supply_assignee_of.get(sc)
        s_has_pi = bool(_hide_fx_only(_forms_of(FormKind.PI, side=sc)))
        s_has_to = bool(_hide_fx_only(_forms_of(FormKind.TO, side=sc)))
        side_holder = case.side_holder(sc)
        side_status = case.side_status(sc)
        side_terminal = side_status in CaseStatus.TERMINAL
        side_at_supply = (side_holder == Unit.SUPPLY)
        side_at_tech = (side_holder == Unit.TECHNICAL)
        # A side that has been sent to the client (CLOSED) or marked final
        # (FINAL_APPROVED) is no longer "at Commercial" for ordinary work actions
        # (submit / return / send-to-client). Only Final Approved / Burned /
        # Final Closed remain, and those are gated separately below.
        side_closed_like = side_status in (CaseStatus.CLOSED, CaseStatus.FINAL_APPROVED)
        side_at_comm = (side_holder == Unit.COMMERCIAL
                        and not side_terminal and not side_closed_like)
        tech_assignee_side = (case.technical_internal_assignee_id if sc == Side.INTERNAL
                              else case.technical_external_assignee_id if sc == Side.EXTERNAL else None)
        if case.is_split:
            owns = services.can_act_on_side(
                case, request.user, sc, role=role, work_user=seat_user)
            # Supply
            can_pi_side = is_supply and owns and side_at_supply
            can_assign_side = (is_supply and is_mgr and not s_assignee
                               and side_at_supply and not s_has_pi)
            # Per-side PI-remark block: this side's current PI carrying any
            # filled remark cannot be forwarded to commercial. Return-to-technical
            # stays available so supply can hand it back.
            side_pi_blocked = services._pi_blocks_commercial(case, sc)
            can_send_side = (is_supply and owns and side_at_supply and s_has_pi
                             and not side_pi_blocked)
            can_return_side = is_supply and owns and side_at_supply
            can_cannot_side = is_supply and owns and side_at_supply
            # Technical
            can_build_to_side = is_tech and owns and side_at_tech
            # Direction: a side that is RETURNED_TO_TECHNICAL came back from
            # Supply (so Technical forwards to Commercial and returns to Supply);
            # a side that is WITH_TECHNICAL came from Commercial (so Technical
            # forwards to Supply and returns to Commercial). This mirrors the
            # non-split rule so each side behaves exactly like a standalone case.
            side_from_supply = (side_status == CaseStatus.RETURNED_TO_TECHNICAL)
            side_to_blocked = services._to_blocks_supply(case, sc)
            side_to_blocked_issues = services._to_has_technical_problems(case, sc)
            side_to_blocked_brand = bool(services._to_rows_without_brand(case, sc))
            side_pi_blocked_tech = services._pi_blocks_commercial(case, sc)
            if side_from_supply:
                # Forward target = Commercial (blocked while PI has any remark,
                # since Technical can't edit the remark — return to Supply instead).
                can_send_supply_side = False
                can_submit_comm_side = (is_tech and owns and side_at_tech and s_has_to
                                        and not side_pi_blocked_tech)
                can_return_comm_side = False
                can_return_supply_from_tech = (is_tech and owns and side_at_tech
                                               and not side_to_blocked)
            else:
                # Forward target = Supply (blocked while any TO problem flag set).
                can_send_supply_side = (is_tech and owns and side_at_tech and s_has_to
                                        and not side_to_blocked)
                can_submit_comm_side = False
                can_return_comm_side = is_tech and owns and side_at_tech
                can_return_supply_from_tech = False
            can_tech_assign_side = False  # technical uses ONE assign (see below)
            # Commercial
            # Send-to-client for a side: TO (+ PI when needed) at the inquiry
            # version, OR a Technical-Problem TO (no Proforma required).
            can_close_side = (is_comm and owns and side_at_comm
                              and side_status != CaseStatus.UNSUPPLIABLE
                              and services._can_send_to_client(case, sc))
            can_cancel_side = (is_comm and owns and side_at_comm
                               and side_status != CaseStatus.PENDING_CANCEL)
            currency_only = services.is_currency_conversion_only(case, sc)
            # Currency-conversion-only reopen: no routing to Technical / Supply.
            can_submit_tech_side = (is_comm and owns and side_at_comm
                                    and not currency_only
                                    and side_status != CaseStatus.PENDING_CANCEL)
            # Return to Supply only when this side arrived FROM Supply.
            can_return_supply_side = (is_comm and owns and side_at_comm and s_has_pi
                                      and side_status == CaseStatus.WITH_COMMERCIAL
                                      and not currency_only)
            can_finalize_side = (is_comm and owns
                                 and case.side_status(sc) == CaseStatus.CLOSED)
            can_final_close_side = (is_comm and owns
                                    and case.side_status(sc) == CaseStatus.FINAL_APPROVED)
            can_burn_side = (is_comm and owns
                             and case.side_status(sc) == CaseStatus.CLOSED)
            can_approve_cancel_side = (
                bool(is_comm and is_mgr)
                and side_status == CaseStatus.PENDING_CANCEL
            )
        else:
            can_pi_side = is_supply and ("build_pi" in actions) and (
                s_assignee == seat_id or (is_mgr and not s_assignee))
            can_assign_side = (is_supply and is_mgr and ("assign" in actions)
                               and not s_assignee and not s_has_pi)
            can_send_side = can_return_side = can_cannot_side = False
            can_build_to_side = ("build_to" in actions)
            can_submit_comm_side = False
            can_return_supply_from_tech = False
            can_send_supply_side = False
            can_return_comm_side = can_tech_assign_side = False
            can_submit_tech_side = can_close_side = can_cancel_side = False
            can_return_supply_side = False
            can_finalize_side = False
            can_final_close_side = False
            can_burn_side = False
            can_approve_cancel_side = False
        inq_cur = s_forms[FormKind.INQUIRY]
        inq_sent = bool(inq_cur and inq_cur.sent)
        if case.is_split:
            can_edit_inq = is_comm and owns and side_at_comm and not inq_sent
            # A new inquiry version for a side is allowed once that side has been
            # sent to the client (CLOSED), or for the freshly-converted two-stage
            # side (it carries a copied inquiry to revise via New version).
            can_new_inq = services.can_new_inquiry_version(
                case, request.user, sc, role=role, work_user=seat_user)
        else:
            can_edit_inq = ("edit" in actions)
            can_new_inq = ("new_inquiry_version" in actions)
        s_events = [e for e in events if (e.side == sc or not e.side)]
        s_arrival, s_assign = _side_notes(events_recent, sc, side_holder)
        sides_data.append({
            "code": sc,
            "label": Side.LABELS.get(sc, "Items"),
            "forms": s_forms,
            "version_lists": s_vlists,
            "to_mode": _form_mode(FormKind.TO, sc),
            "pi_mode": _form_mode(FormKind.PI, sc),
            "can_pi": can_pi_side,
            "can_assign": can_assign_side,
            "can_send": can_send_side,
            "can_return": can_return_side,
            "can_cannot": can_cannot_side,
            "can_build_to": can_build_to_side,
            "can_send_supply": can_send_supply_side,
            # Only when this side would forward to Supply (not returned-from-Supply).
            "to_blocked": (case.is_split and is_tech and owns and side_at_tech
                           and s_has_to and not side_from_supply
                           and services._to_blocks_supply(case, sc)),
            "to_blocked_issues": (case.is_split and is_tech and owns and side_at_tech
                                  and s_has_to and not side_from_supply
                                  and side_to_blocked_issues),
            "to_blocked_brand": (case.is_split and is_tech and owns and side_at_tech
                                 and s_has_to and not side_from_supply
                                 and side_to_blocked_brand),
            "pi_blocked": (case.is_split and s_has_pi
                           and services._pi_blocks_commercial(case, sc)
                           and ((is_supply and owns and side_at_supply)
                                or (is_tech and owns and side_at_tech
                                    and side_status == CaseStatus.RETURNED_TO_TECHNICAL))),
            "can_return_comm": can_return_comm_side,
            "can_submit_comm": can_submit_comm_side,
            "can_return_supply_from_tech": can_return_supply_from_tech,
            "can_tech_assign": can_tech_assign_side,
            "can_submit_tech": can_submit_tech_side,
            "can_close": can_close_side,
            "can_cancel": can_cancel_side,
            "can_approve_cancel": can_approve_cancel_side,
            "can_return_supply": can_return_supply_side,
            "can_finalize": can_finalize_side,
            "can_final_close": can_final_close_side,
            "can_burn": can_burn_side,
            "can_edit_inq": can_edit_inq,
            "can_new_inq": can_new_inq,
            "inq_sent": inq_sent,
            "tech_assignee": (case.technical_internal_assignee if sc == Side.INTERNAL
                              else case.technical_external_assignee if sc == Side.EXTERNAL else None),
            "at_supply": side_at_supply,
            "at_tech": side_at_tech,
            "at_comm": side_at_comm,
            "side_status_label": CaseStatus.LABELS.get(case.side_status(sc), case.side_status(sc)),
            "side_status_color": CaseStatus.COLORS.get(case.side_status(sc), "#6b7280"),
            "side_holder": side_holder,
            "assignee": (case.supply_internal_assignee if sc == Side.INTERNAL
                         else case.supply_external_assignee if sc == Side.EXTERNAL else None),
            "events": s_events,
            "arrival_comment": s_arrival,
            "assign_comment": s_assign,
        })

    # Per-side supply expert pools for the assign dropdowns.
    from accounts.constants import SupplyKind
    pool_internal, pool_external = [], []
    if is_supply and is_mgr:
        pool_internal = list(User.objects.filter(
            profile__unit=Unit.SUPPLY, profile__role=Role.EXPERT,
            profile__supply_kind=SupplyKind.INTERNAL, is_active=True).select_related("profile"))
        pool_external = list(User.objects.filter(
            profile__unit=Unit.SUPPLY, profile__role=Role.EXPERT,
            profile__supply_kind=SupplyKind.EXTERNAL, is_active=True).select_related("profile"))
    tech_pool = []
    if is_tech and is_mgr:
        tech_pool = list(User.objects.filter(
            profile__unit=Unit.TECHNICAL, profile__role=Role.EXPERT,
            is_active=True).select_related("profile"))
    for sd in sides_data:
        sd["pool"] = pool_internal if sd["code"] == Side.INTERNAL else (
            pool_external if sd["code"] == Side.EXTERNAL else [])
        sd["tech_pool"] = tech_pool

    # Technical split case: ONE assign control for both sides. It disappears
    # once an expert is assigned OR the manager has built a TO for either side.
    tech_split_can_assign = bool(
        is_tech and is_mgr and case.is_split
        and (case.side_holder(Side.INTERNAL) == Unit.TECHNICAL
             or case.side_holder(Side.EXTERNAL) == Unit.TECHNICAL)
        and not case.technical_assignee_id
        and not _forms_of(FormKind.TO))

    # Strict per-side visibility: a supply EXPERT only ever sees the side they
    # were assigned (never the other side, never the combined timeline).
    hide_combined = False
    if is_supply and not is_mgr:
        my_codes = []
        if case.supply_internal_assignee_id == seat_id:
            my_codes.append(Side.INTERNAL)
        if case.supply_external_assignee_id == seat_id:
            my_codes.append(Side.EXTERNAL)
        if my_codes:
            sides_data = [sd for sd in sides_data if sd["code"] in my_codes]
        hide_combined = True
    multi_side = len(sides_data) > 1

    # Full case-information editing is only meaningful on a never-submitted draft.
    _cur_inq = case.current_form(FormKind.INQUIRY)
    # Form versions are 1-based, so the very first inquiry is v01. A "fresh draft"
    # is that first version while the case is still a draft.
    is_fresh_draft = case.status == CaseStatus.DRAFT and (_cur_inq is None or _cur_inq.version <= 1)
    if case.is_split:
        is_fresh_draft = is_fresh_draft and all(
            case.side_status(sc) == CaseStatus.DRAFT for sc in case.sides)

    from django.conf import settings as _dj_settings

    context = {
        "case": case,
        "actions": actions,
        "line_items": case.line_items.all(),
        "forms_by_kind": forms_by_kind,
        "versions": versions,
        "version_lists": version_lists,
        "doc_kinds": DocKind.CHOICES,
        "offer_types": OfferType.CHOICES,
        "events": events,
        "is_admin_view": is_admin_view,
        "lifecycle_report": _case_lifecycle_report(case) if is_admin_view else None,
        "comment_form": CommentForm(),
        "assignees": assignees,
        "arrival_comment": arrival_comment,
        "assign_comment": assign_comment,
        "to_author": to_author,
        "to_mode": to_mode,
        "pi_mode": pi_mode,
        "is_fresh_draft": is_fresh_draft,
        "sides_data": sides_data,
        "tech_split_can_assign": tech_split_can_assign,
        "tech_pool": tech_pool,
        "is_split": case.is_split,
        "status_rows": services.detail_status_rows(case, request.user),
        "multi_side": multi_side,
        "hide_combined": hide_combined,
        "combined_events": events,
        "FormKind": FormKind,
        # Warn when TO has Technical Problem rows — no Supply routing allowed.
        "to_blocked_by_issues": (
            not case.is_split
            and case.holder_unit == Unit.TECHNICAL
            and bool(_forms_of(FormKind.TO))
            and services._to_has_technical_problems(case)
            and viewer_unit == Unit.TECHNICAL
        ),
        # Warn when any active TO row still has an empty BRAND.
        "to_blocked_by_brand": (
            not case.is_split
            and case.holder_unit == Unit.TECHNICAL
            and bool(_forms_of(FormKind.TO))
            and bool(services._to_rows_without_brand(case))
            and not services._to_has_technical_problems(case)
            and viewer_unit == Unit.TECHNICAL
        ),
        # Warn when the current Proforma has remark rows blocking Submit to
        # Commercial. Shown both to Supply (who must clear the remark) and to
        # Technical when the case came back from Supply still carrying a remark
        # (Technical can't edit it, so it can only return to Supply).
        "pi_blocked_by_remark": (
            not case.is_split
            and bool(_forms_of(FormKind.PI))
            and services._pi_blocks_commercial(case)
            and (
                (case.holder_unit == Unit.SUPPLY and viewer_unit == Unit.SUPPLY)
                or
                (case.holder_unit == Unit.TECHNICAL and viewer_unit == Unit.TECHNICAL)
            )
        ),
        "is_comm": is_comm,
        "is_tech": is_tech,
        "is_supply": is_supply,
        "viewer_unit": viewer_unit or "",
        # Admin-only export audit timeline (who exported what, and when).
        "export_logs": (list(case.export_logs.select_related("actor").all()[:300])
                        if is_admin_view else None),
        "currency_logs": (list(case.currency_logs.select_related("actor").all()[:300])
                          if is_admin_view else None),
        "vat_percent": _vat_percent_value(),
        "require_ftco_code": bool(getattr(_dj_settings, "REQUIRE_FTCO_CODE_TO_SUPPLY", True)),
    }
    return render(request, "cases/case_detail.html", context)


def _event_visible_to(event, unit) -> bool:
    """An event is visible to a unit when that unit performed it, or when the
    event is a handoff into or out of that unit. System events are shown to all.

    Prefer frozen ``from_unit`` / ``to_unit`` over the actor's live profile so
    history stays readable after seat reassignment / Delegate.
    """
    if unit is None:
        return True
    if event.from_unit == unit or event.to_unit == unit:
        return True
    # Always show DELEGATE on the case timeline (frozen ownership transfer).
    if getattr(event, "action", "") == EventAction.DELEGATE:
        return True
    actor = getattr(event, "actor", None)
    actor_unit = getattr(getattr(actor, "profile", None), "unit", None)
    if actor_unit is None:
        return True  # system / unattributed events
    return actor_unit == unit


def _newver_context(case, rows, side, offer_type, price_type, seeded=False,
                    currency_conversion=False, update_price=False):
    """Build the edit_items render context for "New version" mode.

    ``rows`` is a list of dicts with keys client_row/description/size/quantity/
    unit (seed format). Canonical inquiry keys (#/Description/…) are also
    accepted so a refused save can re-render without wiping the grid.
    """
    line_items = []
    for idx, r in enumerate(rows or [], start=1):
        r = r or {}
        cr = r.get("client_row", r.get("#", idx))
        line_items.append(SimpleNamespace(
            client_row=cr if cr not in (None, "") else idx,
            row_no=idx,
            description=r.get("description", r.get("Description", "")),
            size=r.get("size", r.get("Size", "")),
            quantity=r.get("quantity", r.get("Qty", "")),
            unit=r.get("unit", r.get("Unit", "")),
            deleted=str(r.get("deleted", r.get("_deleted", "")) or "") == "1",
            added=str(r.get("added", r.get("_added", "")) or "") == "1",
            comment=str(r.get("comment", r.get("_comm_comment", "")) or "").strip(),
        ))
    return {
        "case": case,
        "line_items": line_items,
        "editable_meta": False,      # New version: only the table (and deadline) change.
        "editable_contacts": False,
        "show_items": True,
        "show_deadline": True,
        "edit_side": side,
        "newver_mode": True,
        "nv_offer": offer_type,
        "nv_price": price_type,
        "nv_currency": bool(currency_conversion),
        "nv_update_price": bool(update_price),
        "doc_kinds": DocKind.CHOICES, "offer_types": OfferType.CHOICES,
        "price_types": PriceType.CHOICES,
        "clients": Client.objects.all().order_by("name"),
    }


@login_required
def edit_items(request, pk):
    case = get_object_or_404(Case, pk=pk)
    profile = _profile(request.user)
    from people.role_nav import work_context
    ctx = work_context(request)
    side = (request.GET.get("side", "") or request.POST.get("edit_side", "")).strip()
    is_comm = bool(
        (ctx.role and ctx.role.unit == Unit.COMMERCIAL)
        or (profile and profile.unit == Unit.COMMERCIAL)
    )

    # ------------------------------------------------------------------
    # "New version" mode. Reached from the New-version button (GET) and its
    # own submit (POST). A NEW inquiry version is committed only if the table
    # actually changes (or a two-stage upgrade is requested); otherwise the
    # editor re-renders with an error and the user stays on the page.
    # ------------------------------------------------------------------
    newver = (request.GET.get("newver", "") or request.POST.get("newver", "")).strip() == "1"
    if newver:
        nv_side = side if (case.is_split and side in (Side.INTERNAL, Side.EXTERNAL)) else ""
        if not services.can_new_inquiry_version(case, request.user, nv_side):
            messages.error(request, "A new version can't be started for this case right now.")
            return redirect("cases:case_detail", pk=pk)

        # Upgrade flags chosen on the New-version toggle (carried as params).
        nv_offer = (request.GET.get("offer_type", "") or request.POST.get("offer_type", "")).strip()
        nv_price = (request.GET.get("price_type", "") or request.POST.get("price_type", "")).strip()
        nv_currency = (request.GET.get("currency_conversion", "")
                       or request.POST.get("currency_conversion", "")).strip() in ("1", "true", "yes", "on")
        nv_update_price = (request.GET.get("update_price", "")
                           or request.POST.get("update_price", "")).strip() in ("1", "true", "yes", "on")
        # Unit conversion / Update price only for TO & PI cases.
        if case.offer_type != OfferType.TO_PI:
            nv_currency = False
            nv_update_price = False

        # The version we are branching from (this side's current inquiry).
        nv_cur = case.current_form(FormKind.INQUIRY, nv_side) or \
                 (case.current_form(FormKind.INQUIRY) if not nv_side else None)
        nv_rows = list(nv_cur.table or []) if nv_cur else []

        if request.method == "POST":
            try:
                rows = json.loads(request.POST.get("rows", "[]"))
            except ValueError:
                rows = []
            inq_errs = validate_inquiry_rows(rows)
            if inq_errs:
                for err in inq_errs:
                    messages.error(request, err)
                return render(request, "cases/edit_items.html",
                              _newver_context(case,
                                              services.apply_inquiry_row_marks_vs_v00(
                                                  [
                                                      {
                                                          "client_row": r.get("client_row", ""),
                                                          "description": r.get("description", ""),
                                                          "size": r.get("size", ""),
                                                          "quantity": r.get("quantity", ""),
                                                          "unit": r.get("unit", ""),
                                                          "_deleted": "1" if str(r.get("deleted", "") or "") == "1" else "",
                                                          "_added": "1" if str(r.get("added", "") or "") == "1" else "",
                                                          "_comm_comment": str(r.get("comment", "") or "").strip(),
                                                      }
                                                      for r in rows
                                                  ],
                                                  services.v00_client_row_set(case, nv_side)),
                                              nv_side, nv_offer, nv_price,
                                              currency_conversion=nv_currency,
                                              update_price=nv_update_price))
            # Build the canonical inquiry table from the submitted grid. # is
            # taken from each surviving row's data-client (so deletions leave a
            # visible gap); Item reflows 1..N.
            new_table = []
            for idx, row in enumerate(rows, start=1):
                try:
                    cr = int(str(row.get("client_row", "")).strip() or idx)
                except (ValueError, TypeError):
                    cr = idx
                entry = {
                    "#": cr,
                    "Item": idx,
                    "Description": str(row.get("description", "")).strip(),
                    "Size": str(row.get("size", "")).strip(),
                    "Qty": str(row.get("quantity", "")).strip(),
                    "Unit": str(row.get("unit", "")).strip(),
                }
                if str(row.get("deleted", "") or "") == "1":
                    entry["_deleted"] = "1"
                if str(row.get("added", "") or "") == "1":
                    entry["_added"] = "1"
                note = str(row.get("comment", "") or "").strip()
                if note:
                    entry["_comm_comment"] = note
                new_table.append(entry)
            try:
                version = services.commit_inquiry_version(
                    case, request.user, new_table=new_table, side=nv_side,
                    offer_type=nv_offer, price_type=nv_price,
                    currency_conversion=nv_currency,
                    update_price=nv_update_price,
                )
            except services.InquiryUnchanged:
                # No change + no upgrade -> refuse, stay on the page.
                messages.error(
                    request,
                    "No new version was created: change the table (edit a cell, "
                    "add or delete a row) — or turn on Unit conversion / Update "
                    "price / the two-stage upgrade — before saving.")
                return render(request, "cases/edit_items.html",
                              _newver_context(case,
                                              services.apply_inquiry_row_marks_vs_v00(
                                                  new_table, services.v00_client_row_set(case, nv_side)),
                                              nv_side, nv_offer, nv_price,
                                              currency_conversion=nv_currency,
                                              update_price=nv_update_price))
            messages.success(request, f"New inquiry version {version:02d} saved.")
            if nv_side:
                return redirect(f"{reverse('cases:case_detail', args=[pk])}?side={nv_side}")
            return redirect("cases:case_detail", pk=pk)

        # GET: show the editor seeded with the current version's rows.
        v00_rows = services.v00_client_row_set(case, nv_side)
        seed = []
        for r in nv_rows:
            seed.append({
                "client_row": r.get("#", r.get("client_row", "")),
                "description": r.get("Description", r.get("description", "")),
                "size": r.get("Size", r.get("size", "")),
                "quantity": r.get("Qty", r.get("quantity", "")),
                "unit": r.get("Unit", r.get("unit", "")),
                "_deleted": r.get("_deleted", ""),
                "_added": r.get("_added", ""),
                "_comm_comment": r.get("_comm_comment", ""),
            })
        seed = services.apply_inquiry_row_marks_vs_v00(seed, v00_rows)
        return render(request, "cases/edit_items.html",
                      _newver_context(case, seed, nv_side, nv_offer, nv_price, seeded=True,
                                      currency_conversion=nv_currency,
                                      update_price=nv_update_price))

    # Per-side editing for a combined case: a side can be edited independently
    # whenever it is back with Commercial and its current inquiry is unsent —
    # exactly like a standalone (non-combined) case.
    side_edit = False
    contacts_only = (request.GET.get("contacts", "") or request.POST.get("contacts", "")).strip() == "1"
    actions_now = services.allowed_actions_for_request(case, request)
    can_edit_inquiry = "edit" in actions_now
    can_edit_contacts = "edit_contacts" in actions_now

    # Contacts-only mode (Case information → Edit): never touch the inquiry table.
    if contacts_only:
        if not can_edit_contacts:
            messages.error(request, "Contact fields cannot be edited right now.")
            return redirect("cases:case_detail", pk=pk)
        if request.method == "POST":
            case.client_commercial_expert = request.POST.get("client_commercial_expert", "").strip()
            case.client_commercial_phone = request.POST.get("client_commercial_phone", "").strip()
            case.client_technical_expert = request.POST.get("client_technical_expert", "").strip()
            case.client_technical_phone = request.POST.get("client_technical_phone", "").strip()
            case.save(update_fields=[
                "client_commercial_expert", "client_commercial_phone",
                "client_technical_expert", "client_technical_phone", "updated_at",
            ])
            messages.success(request, "Case contacts updated.")
            return redirect("cases:case_detail", pk=pk)
        return render(request, "cases/edit_items.html", {
            "case": case,
            "line_items": [],
            "editable_meta": False,
            "editable_contacts": True,
            "show_items": False,
            "show_deadline": False,
            "contacts_only": True,
            "edit_side": "",
            "doc_kinds": DocKind.CHOICES, "offer_types": OfferType.CHOICES,
            "price_types": PriceType.CHOICES,
            "clients": Client.objects.all().order_by("name"),
        })

    if case.is_split and side in (Side.INTERNAL, Side.EXTERNAL):
        cur_s = case.current_form(FormKind.INQUIRY, side)
        side_at_comm = (case.side_holder(side) == Unit.COMMERCIAL
                        and case.side_status(side) not in CaseStatus.TERMINAL)
        owns = services.can_act_on_side(case, request.user, side)
        side_edit = bool(is_comm and owns and side_at_comm and not (cur_s and cur_s.sent))
        if not side_edit:
            messages.error(request, "This side can no longer be edited here.")
            return redirect("cases:case_detail", pk=pk)
    else:
        if not can_edit_inquiry:
            messages.error(request, "This case can no longer be edited here.")
            return redirect("cases:case_detail", pk=pk)
        # On a split case, allow editing as long as the CURRENT inquiry version is
        # unsent. Older sent versions are fine — the user just made a fresh
        # (unsent) version and wants to edit/delete rows before submitting. Only
        # block when the current version has already been sent.
        if case.is_split and can_edit_inquiry:
            cur_any = case.current_form(FormKind.INQUIRY, "") or \
                      case.current_form(FormKind.INQUIRY, Side.INTERNAL) or \
                      case.current_form(FormKind.INQUIRY, Side.EXTERNAL)
            if cur_any and cur_any.sent:
                messages.error(request, "A side has already been submitted. Create a new version for the side you want to change.")
                return redirect("cases:case_detail", pk=pk)

    # Full case-information editing is only offered on a brand-new draft that
    # has never been submitted. The moment any inquiry version has been sent
    # (i.e. at least one action was taken), only the deadline may change — for
    # both combined and non-combined cases.
    cur_inq = case.current_form(FormKind.INQUIRY, side if side_edit else "")
    if side_edit:
        # A per-side edit is judged ONLY by that side: a freshly-created side
        # (e.g. the new External stream after a two-stage conversion) that is at
        # Draft/Commercial with an unsent v01 inquiry may edit and delete rows
        # exactly like a brand-new case — independent of the other side's state.
        side_inq_sent = bool(cur_inq and cur_inq.sent)
        is_fresh_draft = (case.side_status(side) == CaseStatus.DRAFT
                          and not side_inq_sent
                          and (cur_inq is None or cur_inq.version <= 1))
    else:
        inquiry_ever_sent = case.forms.filter(kind=FormKind.INQUIRY, sent=True).exists()
        is_fresh_draft = (case.status == CaseStatus.DRAFT and not inquiry_ever_sent
                          and (cur_inq is None or cur_inq.version <= 1))
        if case.is_split:
            is_fresh_draft = is_fresh_draft and all(
                case.side_status(sc) == CaseStatus.DRAFT for sc in case.sides)
    # Deadline may be (re)set while the case is a draft, or for an editable side.
    deadline_editable = (case.status == CaseStatus.DRAFT) or side_edit

    if request.method == "POST":
        # Parse the (optional) deadline. Invalid or past dates are ignored
        # silently (the previous deadline is kept) so a bad date never blocks
        # saving the user's row edits.
        new_deadline = case.deadline
        if deadline_editable:
            deadline_raw = (request.POST.get("deadline", "") or "").strip()
            if deadline_raw:
                dt = _parse_jalali_deadline(deadline_raw)
                if dt is not None:
                    new_deadline = dt
            else:
                new_deadline = None

        rows_json = request.POST.get("rows", "[]")
        try:
            rows = json.loads(rows_json)
        except ValueError:
            rows = []

        inq_errs = validate_inquiry_rows(rows)
        if inq_errs:
            for err in inq_errs:
                messages.error(request, err)
            line_items = []
            for idx, r in enumerate(rows, start=1):
                line_items.append(SimpleNamespace(
                    client_row=r.get("client_row") or idx,
                    row_no=idx,
                    description=r.get("description", ""),
                    size=r.get("size", ""),
                    quantity=r.get("quantity", ""),
                    unit=r.get("unit", ""),
                    deleted=str(r.get("deleted", "") or "") == "1",
                    added=str(r.get("added", "") or "") == "1",
                ))
            return render(request, "cases/edit_items.html", {
                "case": case, "line_items": line_items,
                "editable_meta": is_fresh_draft,
                "editable_contacts": False,
                "show_items": True,
                "show_deadline": deadline_editable,
                "edit_side": side if side_edit else "",
                "doc_kinds": DocKind.CHOICES, "offer_types": OfferType.CHOICES,
                "price_types": PriceType.CHOICES,
                "clients": Client.objects.all().order_by("name"),
            })

        # --- Per-side edit on a split case: write ONLY this side's inquiry ---
        # The global line_items pool is shared by both sides, so a per-side edit
        # must NOT delete/recreate it (that would corrupt the other side). We
        # build this side's inquiry table directly from the submitted rows and
        # store it on the side's current inquiry form, leaving the other side and
        # the shared line_items completely untouched.
        #
        # Fresh draft exception: price type / client / kind may still change, and
        # both sides still share one table — fall through to the whole-case path
        # so meta (including price_type) actually persists and inactive side tabs
        # are removed.
        if side_edit and not is_fresh_draft:
            side_table = []
            for idx, row in enumerate(rows, start=1):
                # An already-progressed side keeps each surviving row's original #
                # so soft-deletes keep the row with a − mark.
                try:
                    cr = int(str(row.get("client_row", "")).strip() or idx)
                except (ValueError, TypeError):
                    cr = idx
                entry = {
                    "#": cr,
                    "Item": idx,
                    "Description": str(row.get("description", "")).strip(),
                    "Size": str(row.get("size", "")).strip(),
                    "Qty": str(row.get("quantity", "")).strip(),
                    "Unit": str(row.get("unit", "")).strip(),
                }
                if str(row.get("deleted", "") or "") == "1":
                    entry["_deleted"] = "1"
                if str(row.get("added", "") or "") == "1":
                    entry["_added"] = "1"
                side_table.append(entry)
            inq_form = case.current_form(FormKind.INQUIRY, side)
            if inq_form is not None:
                inq_form.columns = ["#", "Item", "Description", "Size", "Qty", "Unit"]
                inq_form.table = side_table
                inq_form.save(update_fields=["columns", "table"])
            if deadline_editable:
                case.deadline = new_deadline
                case.save(update_fields=["deadline", "updated_at"])
            messages.success(request, "Case updated.")
            return redirect("cases:case_detail", pk=pk)

        case.line_items.all().delete()
        items = []
        flagged_table = []
        for idx, row in enumerate(rows, start=1):
            # Fresh draft (before the first action): the client row (#) reflows
            # 1..N exactly like Item. After the case has moved (a new version),
            # soft-deleted rows stay in the table with _deleted=1.
            if is_fresh_draft:
                cr = idx
            else:
                try:
                    cr = int(str(row.get("client_row", "")).strip() or idx)
                except (ValueError, TypeError):
                    cr = idx
            items.append(LineItem(
                case=case, row_no=idx, client_row=cr,
                description=str(row.get("description", "")).strip(),
                size=str(row.get("size", "")).strip(),
                unit=str(row.get("unit", "")).strip(),
                quantity=str(row.get("quantity", "")).strip(),
            ))
            entry = {
                "#": cr,
                "Item": idx,
                "Description": str(row.get("description", "")).strip(),
                "Size": str(row.get("size", "")).strip(),
                "Qty": str(row.get("quantity", "")).strip(),
                "Unit": str(row.get("unit", "")).strip(),
            }
            if not is_fresh_draft and str(row.get("deleted", "") or "") == "1":
                entry["_deleted"] = "1"
            if not is_fresh_draft and str(row.get("added", "") or "") == "1":
                entry["_added"] = "1"
            flagged_table.append(entry)
        if items:
            LineItem.objects.bulk_create(items)

        if is_fresh_draft:
            # Fresh draft: every field is editable (not logged).
            new_kind = request.POST.get("kind", "")
            new_offer = request.POST.get("offer_type", "")
            new_client = request.POST.get("client", "")
            new_order = request.POST.get("order_no", "")
            if new_kind in dict(DocKind.CHOICES):
                case.kind = new_kind
            if new_offer in dict(OfferType.CHOICES):
                case.offer_type = new_offer
            if new_client:
                try:
                    case.client = Client.objects.get(pk=new_client)
                except (Client.DoesNotExist, ValueError):
                    pass
            case.order_no = new_order.strip()
            new_price = request.POST.get("price_type", "")
            if new_price in dict(PriceType.CHOICES):
                case.price_type = new_price
            # Keep split machinery + inquiry streams in sync with price_type
            # (e.g. BOTH → Internal removes the External tab on a fresh draft).
            services.sync_fresh_draft_price_type(case)
            case.client_commercial_expert = request.POST.get("client_commercial_expert", "").strip()
            case.client_commercial_phone = request.POST.get("client_commercial_phone", "").strip()
            case.client_technical_expert = request.POST.get("client_technical_expert", "").strip()
            case.client_technical_phone = request.POST.get("client_technical_phone", "").strip()
            case.deadline = new_deadline
            case.doc_no = codes.build_doc_no(
                ym=case.year_month, expert_code=case.expert_code,
                client_code=case.client.code, serial=case.serial,
            )
            case.save()
        elif deadline_editable:
            # Inquiry edit path: only the deadline may change alongside the table.
            case.deadline = new_deadline
            case.save(update_fields=["deadline", "updated_at"])
        # Decide which side streams to (re)write. A per-side edit only touches
        # that side. A whole-case edit writes all sides ONLY while the case is a
        # fresh draft (initial v00, shared by both sides). After that, split
        # sides are independent: a new version on one side must never copy into
        # the other side's stream.
        # Fresh draft (even opened via ?side=) rewrites every active side so a
        # price_type change creates/drops Internal/External streams correctly.
        if side_edit and not is_fresh_draft:
            snap_sides = [side]
        elif case.is_split and not is_fresh_draft:
            # Non-side edit on an already-progressed split case: restrict to the
            # side(s) that are currently editable at Commercial (normally none,
            # but guard anyway so we never overwrite an independent side).
            snap_sides = [sc for sc in case.sides
                          if case.side_holder(sc) == Unit.COMMERCIAL
                          and case.side_status(sc) not in CaseStatus.TERMINAL]
            snap_sides = snap_sides or None
        else:
            snap_sides = None
        services._snapshot_inquiry(case, request.user, sides=snap_sides,
                                   table_override=flagged_table)
        # Inquiry edits are intentionally NOT recorded in any timeline.
        messages.success(request, "Case updated.")
        return redirect("cases:case_detail", pk=pk)

    # Prefer the current inquiry snapshot (keeps soft-delete / add marks).
    edit_seed_side = side if side_edit else (case.primary_side or "")
    inq_for_edit = case.current_form(FormKind.INQUIRY, edit_seed_side) or case.current_form(FormKind.INQUIRY)
    if inq_for_edit and inq_for_edit.table and not is_fresh_draft:
        line_items = []
        for idx, r in enumerate(inq_for_edit.table, start=1):
            r = r or {}
            line_items.append(SimpleNamespace(
                client_row=r.get("#", r.get("client_row", idx)) or idx,
                row_no=idx,
                description=r.get("Description", r.get("description", "")),
                size=r.get("Size", r.get("size", "")),
                quantity=r.get("Qty", r.get("quantity", "")),
                unit=r.get("Unit", r.get("unit", "")),
                deleted=str(r.get("_deleted", "") or "") == "1",
                added=str(r.get("_added", "") or "") == "1",
            ))
    else:
        line_items = list(case.line_items.all())
    return render(request, "cases/edit_items.html", {
        "case": case, "line_items": line_items,
        "editable_meta": is_fresh_draft,
        "editable_contacts": False,
        "show_items": True,
        "show_deadline": deadline_editable,
        "edit_side": side if side_edit else "",
        "doc_kinds": DocKind.CHOICES, "offer_types": OfferType.CHOICES,
        "price_types": PriceType.CHOICES,
        "clients": Client.objects.all().order_by("name"),
    })


# ---------------------------------------------------------------------------
# Transitions
# ---------------------------------------------------------------------------
@login_required
def transition(request, pk):
    case = get_object_or_404(Case, pk=pk)
    if request.method != "POST":
        return redirect("cases:case_detail", pk=pk)

    action = request.POST.get("action", "")
    comment = request.POST.get("comment", "").strip()
    side = request.POST.get("side", "")
    from people.role_nav import work_context
    ctx = work_context(request)
    allowed = services.allowed_actions(
        case, request.user, role=ctx.role, work_user=ctx.seat_user,
    )
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    # Currency conversion is a Commercial-only form edit (not a workflow button).
    if action == "convert_pi_currency":
        try:
            form_id = int(request.POST.get("form_id") or 0)
            services.convert_pi_currency(
                case, request.user,
                form_id=form_id,
                from_unit=request.POST.get("from_unit", ""),
                to_unit=request.POST.get("to_unit", ""),
                side=side,
            )
            if wants_json:
                return JsonResponse({"ok": True})
            messages.success(request, "Proforma currency converted.")
            return redirect("cases:case_detail", pk=pk)
        except Exception as exc:
            logger.exception("Case #%s currency conversion failed", pk)
            if wants_json:
                return JsonResponse({"ok": False, "error": str(exc)}, status=400)
            messages.error(request, f"Currency conversion failed: {exc}")
            return redirect("cases:case_detail", pk=pk)

    if action not in allowed and not services.can_do_side_action(
            case, request.user, action, side, role=ctx.role, work_user=ctx.seat_user):
        messages.error(request, "That action is not available right now.")
        return redirect("cases:case_detail", pk=pk)

    # Cancel / burn / final-close / cannot-supply require a non-empty comment
    # (same rule as the UI ``data-required`` confirm panels).
    if action in ("request_cancel", "burn", "final_close", "cannot_supply") and not comment:
        messages.error(request, "A comment is required for this action.")
        return redirect("cases:case_detail", pk=pk)

    actor = request.user
    try:
        _side = request.POST.get("side", "")
        if action == "submit_to_technical":
            services.submit_to_technical(case, actor, comment, side=_side)
        elif action == "return_to_commercial":
            services.return_to_commercial(case, actor, comment, side=_side)
        elif action == "send_to_supply":
            services.send_to_supply(case, actor, comment, side=_side)
        elif action == "return_to_technical":
            services.return_to_technical(case, actor, comment, side=_side)
        elif action == "send_to_commercial":
            services.send_to_commercial(case, actor, comment, side=_side)
        elif action == "send_to_client" and case.is_split and _side:
            services.close_side(case, actor, _side, comment)
        elif action == "close" and case.is_split and _side:
            services.close_side(case, actor, _side, comment)
        elif action == "request_cancel" and case.is_split and _side:
            services.cancel_side(case, actor, _side, comment)
        elif action == "propose_send":
            proposed = request.POST.get("proposed_action", "")
            services.propose_send(case, actor, proposed, comment)
        elif action == "approve_send":
            services.approve_send(case, actor, comment)
        elif action == "assign":
            assignee_id = request.POST.get("assignee")
            assignee = get_object_or_404(User, pk=assignee_id)
            services.assign(case, actor, assignee, comment=comment,
                            side=request.POST.get("side", ""))
        elif action == "close":
            services.close_case(case, actor, comment)
        elif action == "cannot_supply":
            services.mark_cannot_supply(case, actor, comment, side=request.POST.get("side", ""))
        elif action == "approve_unsuppliable":
            services.approve_unsuppliable(case, actor, comment)
        elif action == "reject_unsuppliable":
            services.reject_unsuppliable(case, actor, comment)
        elif action == "return_to_supply":
            services.return_to_supply(case, actor, comment, side=_side)
        elif action == "finalize" and case.is_split and _side:
            services.finalize_side(case, actor, _side, comment)
        elif action == "finalize":
            services.finalize_case(case, actor, comment)
        elif action == "final_close" and case.is_split and _side:
            services.final_close_side(case, actor, _side, comment)
        elif action == "final_close":
            services.final_close_case(case, actor, comment)
        elif action == "burn" and case.is_split and _side:
            services.burn_side(case, actor, _side, comment)
        elif action == "burn":
            services.burn_case(case, actor, comment)
        elif action == "request_cancel":
            services.request_cancel(case, actor, comment)
        elif action == "approve_cancel":
            services.approve_cancel(case, actor, comment)
        elif action == "reject_cancel":
            services.reject_cancel(case, actor, comment)
        elif action == "new_inquiry_version":
            # The version is NOT created here any more. "New version" simply opens
            # the inquiry editor; a new version is committed there only if the
            # table actually changes (or a two-stage upgrade is requested). Carry
            # the chosen offer/price upgrade through as query params.
            params = {"newver": "1"}
            ot = request.POST.get("offer_type", "")
            pt = request.POST.get("price_type", "")
            if ot:
                params["offer_type"] = ot
            if pt:
                params["price_type"] = pt
            if case.is_split and _side:
                params["side"] = _side
            qs = urlencode(params)
            messages.info(request, "Edit the items — a new version is saved only if you change the table.")
            return redirect(f"{reverse('cases:edit_items', args=[pk])}?{qs}")
        elif action == "upgrade_two_stage":
            services.upgrade_two_stage(case, actor, comment)
            messages.success(request, "Converted to Internal & External two stage.")
            return redirect("cases:case_detail", pk=pk)
        elif action == "comment":
            if comment:
                ev = services.add_comment(case, actor, comment,
                                          side=request.POST.get("side", ""))
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse({
                        "ok": True,
                        "actor": ev.actor_display_name or (actor.get_full_name() or actor.username),
                        "substitute": bool(ev.actor_is_substitute),
                        "comment": ev.comment,
                    })
            elif request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"ok": False, "error": "Empty comment."}, status=400)
        else:
            messages.error(request, "Unknown action.")
            return redirect("cases:case_detail", pk=pk)
    except Exception as exc:  # pragma: no cover - defensive
        # Surface a short message to the user, but also log the full traceback so
        # a failed workflow action is never silently lost from the server logs.
        logger.exception("Case #%s action failed", pk)
        messages.error(request, f"Action failed: {exc}")
        return redirect("cases:case_detail", pk=pk)

    messages.success(request, "Done.")
    return redirect("cases:case_detail", pk=pk)


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------
def _resolve_export_form(case, form_kind, side: str = "", version=None):
    kind = (form_kind or "").upper()
    side = (side or "").strip()
    # An explicit version wins: the export follows the version the user selected
    # (e.g. v00 vs v03), not always the latest. Falls back to current when the
    # version is missing/unknown.
    if version is not None and str(version).strip() != "":
        try:
            vnum = int(version)
        except (TypeError, ValueError):
            vnum = None
        if vnum is not None:
            qs = CaseForm.objects.filter(case=case, kind=kind, version=vnum)
            if side:
                qs = qs.filter(side=side)
            f = qs.order_by("-id").first()
            if f is not None:
                return f
    if side:
        return case.current_form(kind, side)
    return case.current_form(kind)


def _resolve_export_form_for_viewer(case, form_kind, side: str, version, profile):
    """Resolve an exportable form the viewer is allowed to see.

    Owners/admin get the normal current (or explicit version). Other units get
    the latest snapshot published to them when the live current is not yet
    visible (e.g. Technical edited after sending elsewhere).
    """
    form = _resolve_export_form(case, form_kind, side, version=version)
    if form is None:
        return None
    if _user_can_see_exported_form(profile, form):
        return form
    # Explicit version that is not published to this viewer → deny.
    if version is not None and str(version).strip() != "":
        return None
    if profile is None:
        return None
    if profile.is_admin or profile.is_general_manager:
        return form
    owner = _FORM_OWNER_UNIT.get((form_kind or "").upper()) or _FORM_OWNER_UNIT.get(
        getattr(form, "kind", ""))
    if profile.unit == owner:
        return form
    kind = (form_kind or getattr(form, "kind", "") or "").upper()
    qs = CaseForm.objects.filter(case=case, kind=kind)
    if side:
        qs = qs.filter(side=side)
    elif getattr(form, "side", None):
        qs = qs.filter(side=form.side)
    visible = [
        f for f in qs.order_by("version", "id")
        if services.form_published_to_unit(f, profile.unit)
    ]
    return visible[-1] if visible else None


def _parse_terms_post(request, kind: str = "PI") -> dict:
    from .export_data import default_terms_for

    defaults = default_terms_for(kind)
    cats = []
    for i, default in enumerate(defaults["categories"]):
        items_en = [
            ln.strip() for ln in (request.POST.get(f"cat_{i}_items_en") or "").splitlines()
            if ln.strip()
        ]
        items_fa = [
            ln.strip() for ln in (request.POST.get(f"cat_{i}_items_fa") or "").splitlines()
            if ln.strip()
        ]
        cats.append({
            "title_en": request.POST.get(f"cat_{i}_title_en") or default["title_en"],
            "title_fa": request.POST.get(f"cat_{i}_title_fa") or default["title_fa"],
            "full": bool(default.get("full")),
            "items_en": items_en or list(default["items_en"]),
            "items_fa": items_fa or list(default["items_fa"]),
        })
    return {
        "intro_en": request.POST.get("intro_en") or defaults["intro_en"],
        "intro_fa": request.POST.get("intro_fa") or defaults["intro_fa"],
        "categories": cats,
    }


_EXPORT_FMT_LABELS = {
    "xlsx": "Excel", "grouped": "Grouped Excel", "pdf": "PDF", "html": "Print view",
}


def _can_export(profile, form_kind, fmt) -> bool:
    """Per-unit export permissions.

    * Commercial  : every export EXCEPT the supply grouped Excel.
    * Technical   : TO-form exports only (never grouped, never PI).
    * Supply      : the grouped Excel only.
    * Admin / GM  : everything.
    """
    if profile is None:
        return False
    if profile.is_admin or profile.is_general_manager:
        return True
    kind = (form_kind or "").upper()
    fmt = (fmt or "").lower()
    unit = profile.unit
    if unit == Unit.COMMERCIAL:
        return fmt != "grouped"
    if unit == Unit.TECHNICAL:
        return kind == "TO" and fmt != "grouped"
    if unit == Unit.SUPPLY:
        return fmt in ("grouped", "xlsx")
    return False


_FORM_OWNER_UNIT = {
    FormKind.INQUIRY: Unit.COMMERCIAL,
    FormKind.TO: Unit.TECHNICAL,
    FormKind.PI: Unit.SUPPLY,
    "INQUIRY": Unit.COMMERCIAL,
    "TO": Unit.TECHNICAL,
    "PI": Unit.SUPPLY,
}


def _user_can_see_exported_form(profile, form) -> bool:
    """Same recipient rule as case-detail tabs: owner/admin always; others only
    if this snapshot was published to their unit on a handoff."""
    if profile is None or form is None:
        return False
    if profile.is_admin or profile.is_general_manager:
        return True
    owner = _FORM_OWNER_UNIT.get(form.kind)
    if profile.unit == owner:
        return True
    return services.form_published_to_unit(form, profile.unit)


def _log_export(case, form, request, fmt: str, side: str = "") -> None:
    """Record an export for the admin-only export audit timeline."""
    try:
        from .models import CaseExportLog
        kind = getattr(form, "kind", "") or ""
        label = "%s %s" % (kind, _EXPORT_FMT_LABELS.get((fmt or "").lower(), (fmt or "").upper()))
        CaseExportLog.objects.create(
            case=case,
            actor=request.user if getattr(request.user, "is_authenticated", False) else None,
            form_kind=kind,
            form_version=getattr(form, "version", None),
            side=(side or getattr(form, "side", "") or ""),
            fmt=(fmt or "").lower(),
            label=label.strip(),
        )
    except Exception:
        logger.exception("Failed to log export for case #%s", getattr(case, "pk", "?"))



@login_required
def log_currency_conversion(request, pk):
    """AJAX: record a PI unit conversion for the admin/GM Conversion Timeline."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST required"}, status=405)
    case = get_object_or_404(Case, pk=pk)
    profile = _profile(request.user)
    if profile is None:
        return JsonResponse({"ok": False, "error": "Unauthorized"}, status=403)
    # Supply / Commercial / admin / GM may log conversions they perform.
    if not (profile.is_admin or profile.is_general_manager
            or profile.unit in (Unit.SUPPLY, Unit.COMMERCIAL)):
        return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)
    from_unit = (request.POST.get("from_unit") or "").strip()
    to_unit = (request.POST.get("to_unit") or "").strip()
    rate = request.POST.get("rate", "")
    side = (request.POST.get("side") or "").strip()
    reset = (request.POST.get("reset") or "").strip() in ("1", "true", "yes")
    form = case.current_form(FormKind.PI, side) if side else case.current_form(FormKind.PI)
    try:
        services.log_currency_conversion(
            case, request.user,
            from_code=from_unit, to_code=to_unit, rate=rate,
            side=side or getattr(form, "side", "") or "",
            form_kind=FormKind.PI,
            form_version=getattr(form, "version", None),
            source="tool",
            reset=reset,
        )
    except Exception as exc:
        logger.exception("Currency conversion log failed for case #%s", pk)
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)
    return JsonResponse({"ok": True})


@login_required
def export_form(request, pk, form_kind, fmt):
    case = get_object_or_404(Case.objects.select_related("client"), pk=pk)
    profile = _profile(request.user)
    if profile is None:
        return redirect("accounts:login")

    # A user must be able to see this specific case before any export of it is
    # considered — previously this route only checked the unit-level rule
    # below ("does your unit do this kind of export"), which let anyone in the
    # right unit download any case's documents by guessing the case id.
    if not services.user_can_view_case(case, request.user):
        messages.error(request, "You do not have access to this case.")
        return redirect("cases:inbox")

    if not _can_export(profile, form_kind, fmt):
        messages.error(request, "You don't have permission to use this export.")
        return redirect("cases:case_detail", pk=pk)

    side = (request.GET.get("side") or "").strip()
    version = request.GET.get("v")
    form = _resolve_export_form_for_viewer(
        case, form_kind, side, version, profile)
    if form is None:
        messages.error(request, "There is no such form to export yet.")
        return redirect("cases:case_detail", pk=pk)

    # PDF / Print view (HTML): first show the editable bilingual Terms page.
    if fmt in ("pdf", "html") and request.method == "GET":
        from .export_data import (
            client_name_only, default_terms_for, doc_no_export, form_date_jalali,
        )
        confirm_name = (
            "cases:export_form_pdf_confirm" if fmt == "pdf"
            else "cases:export_form_html_confirm"
        )
        confirm_url = reverse(confirm_name, args=[pk, form.kind])
        _cparams = {}
        if side or form.side:
            _cparams["side"] = side or form.side
        if version:
            _cparams["v"] = version
        if _cparams:
            confirm_url = f"{confirm_url}?{urlencode(_cparams)}"
        return render(request, "cases/export/terms_editor.html", {
            "case": case,
            "form": form,
            "side": side or form.side,
            "terms": default_terms_for(form.kind),
            "doc_no": doc_no_export(case, form),
            "client": client_name_only(case, form),
            "form_date": form_date_jalali(form),
            "confirm_url": confirm_url,
            "export_fmt": fmt,
        })

    try:
        if fmt == "xlsx":
            # Supply must not see employer identity (CLIENT / PROJECT) on Excel.
            hide_identity = bool(profile and profile.unit == Unit.SUPPLY
                                 and not profile.is_admin
                                 and not profile.is_general_manager)
            content, filename = exports.export_form_excel(
                case, form,
                hide_client=hide_identity,
                hide_project=hide_identity,
            )
            _log_export(case, form, request, fmt, side)
            return _file_response(content, filename,
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if fmt == "grouped":
            content, filename = exports.export_supply_grouped_excel(case, form)
            _log_export(case, form, request, fmt, side)
            return _file_response(content, filename,
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if fmt == "html":
            # Reached only via POST (terms confirm) or legacy callers.
            from .pdf_export import render_print_view_html
            terms = _parse_terms_post(request, form.kind) if request.method == "POST" else None
            html_out, filename = render_print_view_html(case, form, terms=terms)
            _log_export(case, form, request, fmt, side)
            resp = HttpResponse(html_out, content_type="text/html; charset=utf-8")
            resp["Content-Disposition"] = f'inline; filename="{filename}"'
            return resp
        if fmt == "pdf":
            terms = _parse_terms_post(request, form.kind) if request.method == "POST" else None
            content, filename = exports.export_form_pdf(case, form, terms=terms)
            _log_export(case, form, request, fmt, side)
            return _file_response(content, filename, "application/pdf")
    except Exception as exc:  # never surface a 500 to the user for an export
        logger.exception("Export failed for case #%s (%s/%s)", pk, form_kind, fmt)
        messages.error(request, f"Could not generate the {fmt.upper()} export: {exc}")
        return redirect("cases:case_detail", pk=pk)

    messages.error(request, "Unknown export format.")
    return redirect("cases:case_detail", pk=pk)


def _is_ajax(request) -> bool:
    return (request.headers.get("X-Requested-With") or "").lower() == "xmlhttprequest"


@login_required
def export_form_pdf_confirm(request, pk, form_kind):
    """POST from the Terms editor → generate PDF with the edited texts."""
    if request.method != "POST":
        side = (request.GET.get("side") or "").strip()
        url = reverse("cases:export_form", args=[pk, form_kind, "pdf"])
        params = {}
        if side:
            params["side"] = side
        _v = request.GET.get("v")
        if _v:
            params["v"] = _v
        if params:
            url = f"{url}?{urlencode(params)}"
        return redirect(url)

    case = get_object_or_404(Case.objects.select_related("client"), pk=pk)
    profile = _profile(request.user)
    if profile is None:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "Authentication required."}, status=401)
        return redirect("accounts:login")

    if not services.user_can_view_case(case, request.user):
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "You do not have access to this case."}, status=403)
        messages.error(request, "You do not have access to this case.")
        return redirect("cases:inbox")

    if not _can_export(profile, form_kind, "pdf"):
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "You don't have permission to use this export."}, status=403)
        messages.error(request, "You don't have permission to use this export.")
        return redirect("cases:case_detail", pk=pk)

    side = (request.GET.get("side") or "").strip()
    version = request.GET.get("v")
    form = _resolve_export_form_for_viewer(
        case, form_kind, side, version, profile)
    if form is None:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "There is no such form to export yet."}, status=404)
        messages.error(request, "There is no such form to export yet.")
        return redirect("cases:case_detail", pk=pk)

    try:
        terms = _parse_terms_post(request, form.kind)
        content, filename = exports.export_form_pdf(case, form, terms=terms)
        _log_export(case, form, request, "pdf", side)
        # AJAX: return JSON+base64 so browser extensions (IDM) cannot intercept
        # application/pdf attachment responses and corrupt the download.
        if _is_ajax(request):
            import base64
            return JsonResponse({
                "ok": True,
                "filename": filename,
                "pdf_base64": base64.b64encode(content).decode("ascii"),
            })
        return _file_response(content, filename, "application/pdf")
    except Exception as exc:
        logger.exception("PDF export failed for case #%s (%s)", pk, form_kind)
        if _is_ajax(request):
            return JsonResponse(
                {"ok": False, "error": f"Could not generate the PDF export: {exc}"},
                status=500,
            )
        messages.error(request, f"Could not generate the PDF export: {exc}")
        return redirect("cases:case_detail", pk=pk)


@login_required
def export_form_html_confirm(request, pk, form_kind):
    """POST from the Terms editor → generate Print-view HTML with edited texts."""
    if request.method != "POST":
        side = (request.GET.get("side") or "").strip()
        url = reverse("cases:export_form", args=[pk, form_kind, "html"])
        params = {}
        if side:
            params["side"] = side
        _v = request.GET.get("v")
        if _v:
            params["v"] = _v
        if params:
            url = f"{url}?{urlencode(params)}"
        return redirect(url)

    case = get_object_or_404(Case.objects.select_related("client"), pk=pk)
    profile = _profile(request.user)
    if profile is None:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "Authentication required."}, status=401)
        return redirect("accounts:login")

    if not services.user_can_view_case(case, request.user):
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "You do not have access to this case."}, status=403)
        messages.error(request, "You do not have access to this case.")
        return redirect("cases:inbox")

    if not _can_export(profile, form_kind, "html"):
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "You don't have permission to use this export."}, status=403)
        messages.error(request, "You don't have permission to use this export.")
        return redirect("cases:case_detail", pk=pk)

    side = (request.GET.get("side") or "").strip()
    version = request.GET.get("v")
    form = _resolve_export_form_for_viewer(
        case, form_kind, side, version, profile)
    if form is None:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "There is no such form to export yet."}, status=404)
        messages.error(request, "There is no such form to export yet.")
        return redirect("cases:case_detail", pk=pk)

    try:
        from .pdf_export import render_print_view_html
        terms = _parse_terms_post(request, form.kind)
        html_out, filename = render_print_view_html(case, form, terms=terms)
        _log_export(case, form, request, "html", side)
        if _is_ajax(request):
            import base64
            return JsonResponse({
                "ok": True,
                "filename": filename,
                "html_base64": base64.b64encode(html_out.encode("utf-8")).decode("ascii"),
            })
        resp = HttpResponse(html_out, content_type="text/html; charset=utf-8")
        resp["Content-Disposition"] = f'inline; filename="{filename}"'
        return resp
    except Exception as exc:
        logger.exception("HTML export failed for case #%s (%s)", pk, form_kind)
        if _is_ajax(request):
            return JsonResponse(
                {"ok": False, "error": f"Could not generate the Print view: {exc}"},
                status=500,
            )
        messages.error(request, f"Could not generate the Print view: {exc}")
        return redirect("cases:case_detail", pk=pk)


def _file_response(content: bytes, filename: str, content_type: str) -> HttpResponse:
    from urllib.parse import quote

    safe_name = (filename or "download").replace('"', "").replace("\r", "").replace("\n", "")
    resp = HttpResponse(content, content_type=content_type)
    # ASCII fallback + RFC 5987 so browsers always get a usable name.
    resp["Content-Disposition"] = (
        f'attachment; filename="{safe_name}"; filename*=UTF-8\'\'{quote(safe_name)}'
    )
    # Custom header is always readable by same-origin fetch (unlike some CD configs).
    resp["X-Filename"] = safe_name
    resp["Access-Control-Expose-Headers"] = "Content-Disposition, X-Filename"
    resp["X-Content-Type-Options"] = "nosniff"
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp["Pragma"] = "no-cache"
    return resp


# ---------------------------------------------------------------------------
# Clients (commercial master data) + FX — Commercial manager OR Admin
# ---------------------------------------------------------------------------
def _is_commercial_manager(user) -> bool:
    profile = _profile(user)
    return bool(profile and profile.is_manager and profile.unit == Unit.COMMERCIAL)


def _is_platform_admin(user) -> bool:
    profile = _profile(user)
    return bool(profile and profile.is_admin)


def _can_manage_clients_fx(user) -> bool:
    """Commercial manager or platform admin may open Clients & FX."""
    return _is_commercial_manager(user) or _is_platform_admin(user)


def _deny_clients_fx(request, message: str = "Only the Commercial manager or an Administrator can manage Clients & FX."):
    messages.error(request, message)
    if _is_platform_admin(request.user):
        return redirect("accounts:admin_console")
    return redirect("cases:inbox")


def _parse_rial_price(raw):
    from decimal import Decimal, InvalidOperation
    s = str(raw or "").replace(",", "").replace(" ", "").strip()
    if not s:
        raise ValueError("Enter a Rial price.")
    try:
        n = Decimal(s)
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError("Invalid Rial price.") from exc
    if n < 0:
        raise ValueError("Rial price cannot be negative.")
    return n


@login_required
def master_data_hub(request):
    """Commercial manager / Admin landing: Clients and FX Rates."""
    if not _can_manage_clients_fx(request.user):
        return _deny_clients_fx(request)
    from . import fx_rates as fx
    return render(request, "cases/master_data_hub.html", {
        "can_manage_fx": True,
        "fx_stale": fx.is_rates_stale(),
        "is_admin_md": _is_platform_admin(request.user),
    })


@login_required
def fx_rates(request):
    if not _can_manage_clients_fx(request.user):
        return _deny_clients_fx(request, "Only the Commercial manager or an Administrator can manage FX rates.")
    from . import fx_rates as fx
    rates = fx.list_rates()
    for r in rates:
        r.rial_price_display = fx.format_rial_amount(r.rial_price) if r.rial_price else ""
    return render(request, "cases/fx_rates.html", {
        "rates": rates,
        "catalog": fx.catalog_for_add(),
        "latest_update": fx.latest_update_at(),
        "fx_stale": fx.is_rates_stale(),
    })


@login_required
def fx_rate_add(request):
    if not _can_manage_clients_fx(request.user):
        return _deny_clients_fx(request, "Only the Commercial manager or an Administrator can manage FX rates.")
    if request.method != "POST":
        return redirect("cases:fx_rates")
    from . import fx_rates as fx
    from .models import CurrencyRate
    code = fx.normalize_code(request.POST.get("code", ""))
    if not code or code == "rial":
        messages.error(request, "Choose a currency from the list.")
        return redirect("cases:fx_rates")
    catalog = {c["code"]: c for c in fx.CURRENCY_CATALOG}
    meta = catalog.get(code)
    if meta is None:
        messages.error(request, "Unknown currency.")
        return redirect("cases:fx_rates")
    if CurrencyRate.objects.filter(code=code).exists():
        messages.error(request, f"{code.upper()} is already on the board.")
        return redirect("cases:fx_rates")
    try:
        price = _parse_rial_price(request.POST.get("rial_price"))
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("cases:fx_rates")
    if price <= 0:
        messages.error(request, "Enter a positive Rial price.")
        return redirect("cases:fx_rates")
    CurrencyRate.objects.create(
        code=code,
        name=meta["name"],
        symbol=meta["symbol"],
        rial_price=price,
        is_builtin=False,
        updated_by=request.user,
    )
    messages.success(request, f"{code.upper()} added at {fx.format_rial_amount(price)} Rial.")
    return redirect("cases:fx_rates")


@login_required
def fx_rate_update(request, pk):
    if not _can_manage_clients_fx(request.user):
        return _deny_clients_fx(request, "Only the Commercial manager or an Administrator can manage FX rates.")
    from . import fx_rates as fx
    from .models import CurrencyRate
    row = get_object_or_404(CurrencyRate, pk=pk)
    if request.method != "POST":
        return redirect("cases:fx_rates")
    try:
        price = _parse_rial_price(request.POST.get("rial_price"))
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("cases:fx_rates")
    if price <= 0:
        messages.error(request, "Enter a positive Rial price.")
        return redirect("cases:fx_rates")
    row.rial_price = price
    row.updated_by = request.user
    row.save(update_fields=["rial_price", "updated_by", "updated_at"])
    messages.success(
        request,
        f"{row.code.upper()} updated to {fx.format_rial_amount(price)} Rial.",
    )
    return redirect("cases:fx_rates")


@login_required
def fx_rate_update_all(request):
    """Save every FX board price in one submit and restart the 24h timer.

    Each posted ``rate_<pk>`` is written; ``updated_at`` advances for every
    saved row (same as a single-row Update), so ``is_rates_stale`` resets from
    this moment for the whole board.
    """
    if not _can_manage_clients_fx(request.user):
        return _deny_clients_fx(request, "Only the Commercial manager or an Administrator can manage FX rates.")
    if request.method != "POST":
        return redirect("cases:fx_rates")
    from . import fx_rates as fx
    from .models import CurrencyRate

    rows = list(CurrencyRate.objects.all())
    if not rows:
        messages.error(request, "No currencies on the board yet.")
        return redirect("cases:fx_rates")

    parsed = []
    for row in rows:
        raw = request.POST.get(f"rate_{row.pk}")
        if raw is None:
            messages.error(request, f"Missing price for {row.code.upper()}.")
            return redirect("cases:fx_rates")
        try:
            price = _parse_rial_price(raw)
        except ValueError as exc:
            messages.error(request, f"{row.code.upper()}: {exc}")
            return redirect("cases:fx_rates")
        if price <= 0:
            messages.error(request, f"{row.code.upper()}: enter a positive Rial price.")
            return redirect("cases:fx_rates")
        parsed.append((row, price))

    for row, price in parsed:
        row.rial_price = price
        row.updated_by = request.user
        row.save(update_fields=["rial_price", "updated_by", "updated_at"])

    messages.success(
        request,
        f"Updated {len(parsed)} exchange rate{'s' if len(parsed) != 1 else ''}. "
        f"24-hour timer restarted.",
    )
    return redirect("cases:fx_rates")


@login_required
def fx_rate_delete(request, pk):
    if not _can_manage_clients_fx(request.user):
        return _deny_clients_fx(request, "Only the Commercial manager or an Administrator can manage FX rates.")
    from .models import CurrencyRate
    row = get_object_or_404(CurrencyRate, pk=pk)
    if request.method != "POST":
        return redirect("cases:fx_rates")
    if row.is_builtin or row.code in ("usd", "eur"):
        messages.error(request, "Default currencies (USD / EUR) cannot be removed.")
        return redirect("cases:fx_rates")
    code = row.code.upper()
    row.delete()
    messages.success(request, f"{code} removed from the FX board.")
    return redirect("cases:fx_rates")


@login_required
def fx_rates_api(request):
    """JSON board for PI tool + commercial conversion UIs."""
    profile = _profile(request.user)
    if profile is None:
        return JsonResponse({"ok": False, "error": "Unauthorized"}, status=403)
    unit = profile.unit
    if not (profile.is_admin or profile.is_general_manager
            or unit in (Unit.COMMERCIAL, Unit.SUPPLY)):
        return JsonResponse({"ok": False, "error": "Forbidden"}, status=403)
    from . import fx_rates as fx
    from_unit = request.GET.get("from", "")
    to_unit = request.GET.get("to", "")
    payload = fx.api_payload()
    if from_unit and to_unit:
        try:
            rate, stale = fx.resolve_conversion(from_unit, to_unit)
            payload["from"] = fx.normalize_code(from_unit)
            payload["to"] = fx.normalize_code(to_unit)
            payload["rate"] = rate
            payload["stale"] = stale or payload["stale"]
            payload["convertible"] = (not payload["stale"]) and (
                fx.normalize_code(from_unit) != fx.normalize_code(to_unit)
            )
        except ValueError as exc:
            payload["rate"] = None
            payload["convertible"] = False
            payload["error"] = str(exc)
    return JsonResponse(payload)


@login_required
def client_list(request):
    if not _can_manage_clients_fx(request.user):
        return _deny_clients_fx(request, "Only the Commercial manager or an Administrator can manage clients.")

    query = request.GET.get("q", "").strip()
    profile = _profile(request.user)
    clients = Client.objects.all()
    if query:
        clients = clients.filter(Q(name__icontains=query) | Q(code__icontains=query))

    return render(request, "cases/client_list.html", {
        "clients": clients.order_by("code", "name"),
        "query": query,
        "can_add": bool(profile and profile.can_add_client),
        "can_upload": bool(profile and (profile.can_add_client or _is_platform_admin(request.user))),
        "can_wipe_clients": _is_platform_admin(request.user),
    })


@login_required
def client_add(request):
    profile = _profile(request.user)
    if not (profile and profile.can_add_client):
        return _deny_clients_fx(request, "Only the Commercial manager or an Administrator can add clients.")

    if request.method == "POST":
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save(commit=False)
            client.code = services.next_client_code()
            client.created_by = request.user
            client.save()
            messages.success(request, f"Client added with code {client.code}.")
            return redirect("cases:client_list")
    else:
        form = ClientForm()
    return render(request, "cases/client_form.html", {"form": form})


@login_required
def client_rename(request, pk):
    client = get_object_or_404(Client, pk=pk)
    profile = _profile(request.user)
    if not (profile and profile.can_add_client):
        return _deny_clients_fx(request, "Only the Commercial manager or an Administrator can rename clients.")

    if request.method == "POST":
        form = ClientRenameForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, "Client renamed (the code stays the same).")
            return redirect("cases:client_list")
    else:
        form = ClientRenameForm(instance=client)
    return render(request, "cases/client_form.html", {"form": form, "client": client})


@login_required
def client_upload(request):
    """Bulk client codes from Excel (Commercial manager or Admin)."""
    profile = _profile(request.user)
    if not (profile and profile.can_add_client):
        return _deny_clients_fx(request, "Only the Commercial manager or an Administrator can upload client codes.")

    if request.method == "POST" and request.FILES.get("excel_file"):
        try:
            created, updated, warnings = services.import_clients_from_excel(
                request.FILES["excel_file"], request.user,
            )
        except ValueError as exc:
            messages.error(request, str(exc))
            return render(request, "cases/upload.html", {
                "title": "Upload client codes",
                "hint": "Excel with two columns: Code, Name (.xlsx). Numeric codes become 001, 012, …",
                "action_url": "cases:client_upload",
            })
        except Exception:
            logger.exception("Client Excel upload failed")
            messages.error(
                request,
                "Upload failed unexpectedly. Use a valid .xlsx (Code, Name). "
                "If the file is very large, try again or split it.",
            )
            return render(request, "cases/upload.html", {
                "title": "Upload client codes",
                "hint": "Excel with two columns: Code, Name (.xlsx). Numeric codes become 001, 012, …",
                "action_url": "cases:client_upload",
            })
        messages.success(request, f"Clients imported: {created} new, {updated} updated.")
        for w in (warnings or [])[:12]:
            messages.warning(request, w)
        if warnings and len(warnings) > 12:
            messages.warning(request, f"…and {len(warnings) - 12} more warnings.")
        return redirect("cases:client_list")
    return render(request, "cases/upload.html", {
        "title": "Upload client codes",
        "hint": "Excel with two columns: Code, Name (.xlsx only). Numeric codes are stored as 001, 012, … (4+ digits stay as-is).",
        "action_url": "cases:client_upload",
    })


@login_required
def client_wipe(request):
    """Admin-only: delete every client so a fresh Excel can be uploaded."""
    if not _is_platform_admin(request.user):
        return _deny_clients_fx(request, "Only an Administrator can wipe all clients.")
    if request.method != "POST":
        return redirect("cases:client_list")
    try:
        n = services.wipe_all_clients()
        messages.success(request, f"All clients cleared ({n} removed). You can upload a new Excel now.")
    except ValueError as exc:
        messages.error(request, str(exc))
    return redirect("cases:client_list")


@login_required
def client_delete(request, pk):
    """Admin-only: delete one client if no case uses that client code."""
    if not _is_platform_admin(request.user):
        return _deny_clients_fx(request, "Only an Administrator can delete individual clients.")
    client = get_object_or_404(Client, pk=pk)
    if request.method != "POST":
        return redirect("cases:client_list")
    label = f"{client.code} — {client.name}"
    try:
        services.delete_client_if_unused(client)
        messages.success(request, f"Client deleted: {label}")
    except ValueError as exc:
        messages.error(request, str(exc))
    return redirect("cases:client_list")


@login_required
def client_lookup(request):
    """AJAX search used by the case form (search by code or name)."""
    profile = _profile(request.user)
    if not (profile and (profile.unit == Unit.COMMERCIAL or profile.is_admin)):
        return JsonResponse({"results": []})
    query = request.GET.get("q", "").strip()
    clients = Client.objects.all()
    if query:
        clients = clients.filter(Q(name__icontains=query) | Q(code__icontains=query))
    results = [{"id": c.id, "code": c.code, "name": c.name} for c in clients.order_by("name")[:20]]
    return JsonResponse({"results": results})


# ---------------------------------------------------------------------------
# Expert codes (commercial master data)
# ---------------------------------------------------------------------------
@login_required
def expert_code_list(request):
    profile = _profile(request.user)
    if not (profile and profile.unit == Unit.COMMERCIAL):
        messages.error(request, "Commercial access only.")
        return redirect("cases:inbox")
    return render(request, "cases/expert_code_list.html", {
        "expert_codes": ExpertCode.objects.select_related("user").all(),
        "can_edit": profile.is_manager,
    })


@login_required
def expert_code_add(request):
    profile = _profile(request.user)
    if not (profile and profile.is_manager and profile.unit == Unit.COMMERCIAL):
        messages.error(request, "Only the Commercial manager can edit expert codes.")
        return redirect("cases:expert_code_list")

    if request.method == "POST":
        form = ExpertCodeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Expert code added.")
            return redirect("cases:expert_code_list")
    else:
        form = ExpertCodeForm()
    return render(request, "cases/expert_code_form.html", {"form": form})


@login_required
def expert_code_edit(request, pk):
    profile = _profile(request.user)
    if not (profile and profile.is_manager and profile.unit == Unit.COMMERCIAL):
        messages.error(request, "Only the Commercial manager can edit expert codes.")
        return redirect("cases:expert_code_list")

    expert_code = get_object_or_404(ExpertCode, pk=pk)
    if request.method == "POST":
        form = ExpertCodeForm(request.POST, instance=expert_code)
        if form.is_valid():
            form.save()
            messages.success(request, "Expert code updated.")
            return redirect("cases:expert_code_list")
    else:
        form = ExpertCodeForm(instance=expert_code)
    return render(request, "cases/expert_code_form.html", {"form": form, "expert_code": expert_code})


@login_required
def expert_code_upload(request):
    profile = _profile(request.user)
    if not (profile and profile.is_manager and profile.unit == Unit.COMMERCIAL):
        messages.error(request, "Only the Commercial manager can upload expert codes.")
        return redirect("cases:expert_code_list")

    if request.method == "POST" and request.FILES.get("excel_file"):
        created, updated = _import_two_column(
            request.FILES["excel_file"], ExpertCode, "code", "name", request.user
        )
        messages.success(request, f"Expert codes imported: {created} new, {updated} updated.")
        return redirect("cases:expert_code_list")
    return render(request, "cases/upload.html", {
        "title": "Upload expert codes",
        "hint": "Excel with two columns: Code, Name.",
        "action_url": "cases:expert_code_upload",
    })


def _import_two_column(file_obj, model, code_field, name_field, user) -> tuple[int, int]:
    """Import a simple two-column (code, name) Excel file into a model."""
    import openpyxl

    created = updated = 0
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    for idx, raw in enumerate(ws.iter_rows(values_only=True)):
        cells = list(raw) + [None, None]
        raw_code, raw_name = cells[0], cells[1]
        if model is Client:
            code = services.normalize_client_code(raw_code)
            name = ("" if raw_name is None else str(raw_name).strip())
        else:
            code = ("" if raw_code is None else str(raw_code).strip())
            name = ("" if raw_name is None else str(raw_name).strip())
        if idx == 0 and (code.lower() in {"code", "کد"} or not code):
            continue
        if not code or not name:
            continue
        defaults = {name_field: name}
        if model is Client:
            obj, was_created = model.objects.get_or_create(
                **{code_field: code}, defaults={**defaults, "created_by": user}
            )
        else:
            obj, was_created = model.objects.get_or_create(**{code_field: code}, defaults=defaults)
        if was_created:
            created += 1
        else:
            setattr(obj, name_field, name)
            obj.save(update_fields=[name_field])
            updated += 1
    wb.close()
    return created, updated
