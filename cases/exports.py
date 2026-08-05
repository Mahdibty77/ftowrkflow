"""Export a stored form (Inquiry / TO / PI) to Excel, HTML or PDF.

Output file names follow the business convention, e.g.
``FT-TO-IN-503-102-015-1254-00.xlsx``. For Supply, a grouped Excel export splits
rows by their item Group (one worksheet per group), keeps the original row
numbers, and includes Unit/Total Price with Excel formulas for pricing.

Plain Excel exports include a professional engineering header (FOOLAD TABAR,
Document No., Jalali Date, Client, Order No. / Project Name), a clean data table, footer,
and sheet protection (password ``admin0812``). PI sheets include Subtotal /
VAT / Grand Total. Grouped Excel keeps original item numbers, adds Unit/Total
Price with Excel formulas, unlocks only Unit Price for editing, and adds a
Subtotal SUM row.
"""
from __future__ import annotations

import html
import io
import re

from django.utils import timezone

from .constants import FormKind

PRICE_COLUMN_HINTS = ("unit price", "قیمت فی", "total price", "قیمت کل")
_COL_WIDTHS = {
    "client_no": 11,
    "item": 8,
    "code": 14,
    "desc_client": 28,
    "remark": 18,
    "desc_ftco": 32,
    "size": 10,
    "qty": 8,
    "unit": 8,
    "brand": 12,
    "time": 10,
    "unit_price": 14,
    "total_price": 14,
}

# European engineering palette
_NAVY = "1B3A4B"
_NAVY_DARK = "143040"
_MUTED = "5A6A7A"
_FOOTER = "6B7785"
_LINE = "D0D7DE"
_GRID = "D8DEE6"
_SOFT = "F4F7FA"
_ZEBRA = "F7F9FC"
_WHITE = "FFFFFF"
_INK = "1F2A37"
_TEAL = "0F5C4C"

# Sheet protection password for TO / PI / Grouped Excel exports.
EXCEL_LOCK_PASSWORD = "admin0812"


def _is_price_column(title: str) -> bool:
    t = str(title or "").strip().lower()
    return any(h in t for h in PRICE_COLUMN_HINTS)


def export_name_for(case, form, group_suffix: str = "") -> str:
    from .export_data import export_name_for as _name
    return _name(case, form, group_suffix=group_suffix)


def _export_columns(form) -> list[tuple[str, str]]:
    """Return [(display_title, value_key), ...] — fixed TO/PI export order."""
    from .export_data import export_columns
    return export_columns(form)


def _client_name_only(case, form=None) -> str:
    from .export_data import client_name_only
    return client_name_only(case, form)


def _form_date_jalali(form) -> str:
    from .export_data import form_date_jalali
    return form_date_jalali(form)


def _doc_no(case, form) -> str:
    """Header Doc No. uses the same pattern as the Excel/PDF file name."""
    from .export_data import doc_no_export
    return doc_no_export(case, form)


def _order_no(case, form) -> str:
    from .export_data import order_no
    return order_no(case, form)


def _strip_html(value) -> str:
    s = "" if value is None else str(value)
    if "<" not in s:
        return s
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)
    s = re.sub(r"<[^>]+>", "", s)
    return html.unescape(s).strip()


def _norm_token(value) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _alias_clean(value) -> str:
    """Same cleaning as feature extraction (keep dots)."""
    return re.sub(r"[^a-z0-9آ-ی\.]", "", str(value or "").lower())


def _alias_nodot(value) -> str:
    """Dot-insensitive form so ``grade.b`` matches alias ``gradeb``."""
    return re.sub(r"[^a-z0-9آ-ی]", "", str(value or "").lower())


_REMARK_ALIAS_CACHE: dict[tuple[str, str], list[tuple[str, str, str]]] = {}


def _features_container_for(group: str, type_key: str) -> dict:
    """Build the same features dict the live tool uses for a group/type."""
    from itemcoder.composite_keys import alias_key_matches, get_by_alias
    from itemcoder.regex_patterns import load_json_file
    from itemcoder.resource_paths import json_path

    json_dict = load_json_file(json_path("data.json"))
    group_dict = get_by_alias(json_dict.get("group", {}), group, {}) or {}
    if not isinstance(group_dict, dict):
        return {}

    features: dict = {}
    group_features = group_dict.get("features", {})
    if isinstance(group_features, dict):
        features.update(group_features)

    if type_key:
        for composite_key, composite_val in group_dict.items():
            if isinstance(composite_val, dict) and "features" in composite_val:
                if alias_key_matches(composite_key, type_key) or str(composite_key).strip().lower() == "all_in":
                    features.update(composite_val["features"])

        type_specific = get_by_alias(group_dict, type_key, {})
        if isinstance(type_specific, dict) and "features" in type_specific:
            features.update(type_specific["features"])

        combined = get_by_alias(group_dict, f"{group}-{type_key}", {})
        if isinstance(combined, dict) and "features" in combined:
            features.update(combined["features"])

        type_section = group_dict.get("type", {})
        if isinstance(type_section, dict):
            inner = get_by_alias(type_section, type_key, {})
            if isinstance(inner, dict) and "features" in inner:
                features.update(inner["features"])
    return features


def _phisic_display_token(main_key: str, raw_value: str) -> str:
    """Same compact display as FTCO Description (e.g. ``SCH40``, ``THK: 0.250``)."""
    label = str(main_key or "").strip()
    value = str(raw_value or "").strip()
    if not value:
        return ""
    compact_cores = {"sch", "shc", "cl"}
    if label.lower() in compact_cores:
        return f"{label}{value}".upper()
    return f"{label}: {value}".upper()


def _phisic_alias_pairs_from_key(pat_key: str, pat_values) -> list[tuple[str, str, str]]:
    """Build remark aliases for one phisic JSON key (CSV-backed schedule/thk/…)."""
    from itemcoder.regex_patterns import load_feature_values

    rem = re.sub(r"^M\d+_[A-Z]_?", "", pat_key, count=1)
    main_match = re.search(r"-(.+?)-", rem)
    main_key = main_match.group(1).strip().lower() if main_match else rem.split("_")[-1].strip().lower()
    if not main_key:
        return []

    prefix_section = rem.split("-", 1)[0] if "-" in rem else ""
    suffix_section = rem.split("-", 2)[2] if rem.count("-") >= 2 else ""
    prefixes = [p for p in prefix_section.split("&") if p not in ("", "-")]
    suffixes = [s for s in (suffix_section.split("&") if suffix_section else []) if s not in ("", "-")]
    if not prefixes:
        prefixes = [main_key]

    try:
        vals = load_feature_values(pat_values) or []
    except Exception:
        vals = pat_values if isinstance(pat_values, list) else [pat_values]

    pairs: list[tuple[str, str, str]] = []
    for raw_v in vals:
        v_clean = str(raw_v or "").strip()
        if not v_clean or v_clean.lower() == "null":
            continue
        allow_alone = v_clean.startswith("/")
        raw_core = v_clean.lstrip("/").replace(" ", "").lower()
        display = _phisic_display_token(main_key, raw_core if allow_alone else v_clean)
        if not display:
            continue
        patterns = []
        for p in prefixes:
            if p:
                patterns.append(f"{p.lower()}{raw_core}")
        for s in suffixes:
            if s:
                patterns.append(f"{raw_core}{s.lower()}")
        if allow_alone:
            patterns.append(raw_core)
        # Also accept the final display form itself (SCH40 / sch40).
        patterns.append(display.lower())
        patterns.append(_alias_nodot(display))
        for pat in dict.fromkeys(patterns):
            if not pat:
                continue
            pairs.append((_alias_clean(pat), _alias_nodot(pat), display))
    return pairs


def _remark_alias_pairs(group: str, type_key: str) -> list[tuple[str, str, str]]:
    """Return ``(alias_clean, alias_nodot, canonical)`` sorted longest-first."""
    cache_key = (str(group or "").strip().lower(), str(type_key or "").strip().lower())
    cached = _REMARK_ALIAS_CACHE.get(cache_key)
    if cached is not None:
        return cached

    pairs: list[tuple[str, str, str]] = []
    try:
        from itemcoder.normalizers import parse_feature_dependency_markers, parse_feature_pattern_key
        from itemcoder.regex_patterns import load_feature_values

        features = _features_container_for(cache_key[0], cache_key[1])
        for feature_key, feature_val in (features or {}).items():
            if not isinstance(feature_val, dict):
                continue

            # Physical features use CSV values + prefix/suffix keys — never export
            # the raw JSON key (e.g. ``sch&رده&sch.-sch-``).
            if str(feature_key).startswith("phisic"):
                for pat_key, pat_values in feature_val.items():
                    clean_pat_key, _s, _r = parse_feature_dependency_markers(pat_key)
                    if "null" in clean_pat_key.lower():
                        continue
                    pairs.extend(_phisic_alias_pairs_from_key(clean_pat_key, pat_values))
                continue

            for pat_key, pat_values in feature_val.items():
                clean_pat_key, _suppress, _require = parse_feature_dependency_markers(pat_key)
                if "null" in clean_pat_key.lower():
                    continue
                _mnum, _letter, base_name = parse_feature_pattern_key(clean_pat_key)
                canonical = str(base_name or "").strip()
                if not canonical or canonical.lower() == "null":
                    continue
                synonyms = [canonical]
                try:
                    synonyms.extend(load_feature_values(pat_values) or [])
                except Exception:
                    if isinstance(pat_values, list):
                        synonyms.extend(pat_values)
                for raw in synonyms:
                    raw_s = str(raw or "").strip()
                    if not raw_s or raw_s.lower() == "null":
                        continue
                    pairs.append((_alias_clean(raw_s), _alias_nodot(raw_s), canonical))
    except Exception:
        pairs = []

    # Longest alias first so ``grade.b`` / ``sch40`` win over short fragments.
    pairs.sort(key=lambda item: (-len(item[1] or item[0]), -len(item[0])))
    seen: set[tuple[str, str]] = set()
    unique: list[tuple[str, str, str]] = []
    for ac, an, canon in pairs:
        key = (ac, an)
        if key in seen:
            continue
        seen.add(key)
        unique.append((ac, an, canon))
    _REMARK_ALIAS_CACHE[cache_key] = unique
    return unique


def _match_remark_piece(piece: str, alias_pairs: list[tuple[str, str, str]]) -> str:
    """Map one remark fragment to its data.json canonical display value."""
    pc = _alias_clean(piece)
    pn = _alias_nodot(piece)
    if not pn:
        return ""
    for ac, an, canon in alias_pairs:
        if an and an in pn:
            return canon
        if ac and ac in pc:
            return canon
    return ""


def _remark_from_aliases(remark_text: str, group: str, type_key: str) -> str:
    """Turn raw remark text into canonical tokens in the same order.

    Example: ``grade.b , carbonsteel ,seamless`` → ``Gr.B , C.S , SMLS``.
    """
    raw = str(remark_text or "").strip()
    if not raw:
        return ""
    alias_pairs = _remark_alias_pairs(group, type_key)
    if not alias_pairs:
        return ""

    # Split on commas (and Persian commas) while keeping user order.
    parts = re.split(r"[,،]+", raw)
    mapped: list[str] = []
    for part in parts:
        piece = part.strip()
        if not piece:
            continue
        canon = _match_remark_piece(piece, alias_pairs)
        if canon:
            mapped.append(canon)
    if not mapped:
        # Whole remark as one blob (no commas) — still try a single match.
        canon = _match_remark_piece(raw, alias_pairs)
        return canon
    return " , ".join(mapped)


def _is_export_highlight(style: str, cls: str) -> bool:
    """True for blue/green spans that carry identified feature values."""
    s = f"{style} {cls}".lower()
    return (
        "#001aff" in s
        or "rgb(0, 26, 255)" in s
        or "color:blue" in s
        or "color: blue" in s
        or "color:green" in s
        or "color: green" in s
        or "highlight-color" in s
    )


def _clean_export_segment(value: str) -> str:
    text = re.sub(r"\s+", " ", str(value or ""))
    text = re.sub(r"\s+([,\-/\)])", r" \1", text)
    text = re.sub(r"([\(/\-])\s+", r"\1 ", text)
    return text.strip()


def _segment_from_tokens(tokens: list[dict], revision_norm: str) -> str:
    """Mirror calculation_controls.segmentTextFromNodes for one Final-Text piece."""
    kept = []
    for tok in tokens:
        if tok.get("highlight"):
            text = tok.get("text") or ""
            text_norm = _norm_token(text)
            if not text_norm:
                continue
            if revision_norm and (revision_norm in text_norm or text_norm in revision_norm):
                continue
            kept.append({"highlight": True, "text": text})
        else:
            kept.append({"highlight": False, "text": tok.get("text") or ""})

    highlight_indexes = [i for i, t in enumerate(kept) if t["highlight"]]
    if not highlight_indexes:
        return ""
    if len(highlight_indexes) == 1:
        return _clean_export_segment(kept[highlight_indexes[0]]["text"])

    first, last = highlight_indexes[0], highlight_indexes[-1]
    parts = []
    for i in range(first, last + 1):
        tok = kept[i]
        if tok["highlight"]:
            parts.append(tok["text"])
            continue
        text = tok["text"] or ""
        if text and not re.search(r"[A-Za-z0-9\u0600-\u06FF]", text):
            parts.append(text)
    return _clean_export_segment("".join(parts))


def _remark_from_final_text(final_html: str, revision_text: str = "", fallback: str = "") -> str:
    """Fallback when Final Arranged Text still has highlight HTML spans."""
    from html.parser import HTMLParser

    source = str(final_html or "")
    if "<span" not in source.lower():
        return str(fallback or "").strip()

    revision_norm = _norm_token(revision_text)

    class _Walker(HTMLParser):
        def __init__(self):
            super().__init__(convert_charrefs=True)
            self.nodes: list[dict] = []
            self._stack: list[dict] = []

        def handle_starttag(self, tag, attrs):
            attrs_d = dict(attrs or [])
            self._stack.append({
                "tag": tag.lower(),
                "style": str(attrs_d.get("style") or ""),
                "class": str(attrs_d.get("class") or ""),
                "highlight": False,
                "text_parts": [],
            })
            top = self._stack[-1]
            top["highlight"] = _is_export_highlight(top["style"], top["class"])

        def handle_endtag(self, tag):
            if not self._stack:
                return
            node = self._stack.pop()
            text = "".join(node["text_parts"])
            if node["highlight"]:
                self.nodes.append({"type": "highlight", "text": text})

        def handle_data(self, data):
            if not data:
                return
            if self._stack and self._stack[-1]["highlight"]:
                self._stack[-1]["text_parts"].append(data)
            else:
                self.nodes.append({"type": "text", "text": data})

    walker = _Walker()
    try:
        walker.feed(f"<div>{source}</div>")
        walker.close()
    except Exception:
        return str(fallback or "").strip()

    sep_counts: dict[str, int] = {}
    for node in walker.nodes:
        if node["type"] != "text":
            continue
        value = node["text"]
        trimmed = value.strip()
        simple = trimmed and not re.search(r"[A-Za-z0-9\u0600-\u06FF]", trimmed) and not re.search(r"[/()\[\]{}:]", trimmed)
        blank = (not trimmed) and 0 < len(value) <= 6
        if simple or blank:
            sep_counts[value] = sep_counts.get(value, 0) + 1
    separator = " , "
    if sep_counts:
        separator = max(sep_counts.items(), key=lambda kv: kv[1])[0]

    def _is_sep(node) -> bool:
        if node.get("type") != "text":
            return False
        value = node.get("text") or ""
        if not separator.strip():
            return value == separator
        return value.strip() == separator.strip()

    pieces: list[str] = []
    current: list[dict] = []

    def _flush():
        nonlocal current
        tokens = []
        for n in current:
            if n["type"] == "highlight":
                tokens.append({"highlight": True, "text": n["text"]})
            else:
                tokens.append({"highlight": False, "text": n["text"]})
        piece = _segment_from_tokens(tokens, revision_norm)
        if piece:
            pieces.append(piece)
        current = []

    for node in walker.nodes:
        if _is_sep(node):
            _flush()
        else:
            if (
                not current
                and node.get("type") == "text"
                and separator.strip()
                and (node.get("text") or "").strip().endswith(separator.strip())
                and (node.get("text") or "").strip() != separator.strip()
            ):
                continue
            current.append(node)
    _flush()

    if pieces:
        joined = separator.join(pieces) if separator.strip() else " , ".join(pieces)
        return re.sub(r"\s+", " ", joined).strip()
    return str(fallback or "").strip()


def _excel_cell_value(row: dict, key: str) -> str:
    """Cell value for Excel — Remark uses data.json canonical tokens."""
    if key == "ریمارک":
        raw = _strip_html(_cell(row, "ریمارک"))
        if not raw:
            return ""
        group = str(_cell(row, "Group") or "").strip()
        type_key = str(_cell(row, "Type") or "").strip()
        mapped = _remark_from_aliases(raw, group, type_key)
        if mapped:
            return mapped
        # Older snapshots may still store highlighted Final Arranged Text HTML.
        return _remark_from_final_text(
            _cell(row, "Final Arranged Text"),
            _cell(row, "اصلاحیه"),
            raw,
        )
    return _strip_html(_cell(row, key))


def _sheet_title(form) -> str:
    kind = (form.kind or "").upper()
    if kind == FormKind.TO:
        return "Technical Offer"
    if kind == FormKind.PI:
        return "Proforma Invoice"
    return (form.get_kind_display() or "Form")[:31]


# ---------------------------------------------------------------------------
# openpyxl helpers
# ---------------------------------------------------------------------------
def _font(name="Calibri", size=10, bold=False, color=_INK, italic=False):
    from openpyxl.styles import Font
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def _fill(hex_color):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=hex_color)


def _align(h="left", v="center", wrap=False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(color=_GRID):
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _protect_worksheet(
    ws,
    *,
    unlock_cells: list[tuple[int, int]] | None = None,
    unlock_columns_from: int | None = None,
    used_cols: int = 1,
    max_row: int = 200,
):
    """Lock the sheet with ``EXCEL_LOCK_PASSWORD``.

    ``unlock_cells`` — (row, col) 1-based pairs left editable.
    ``unlock_columns_from`` — 1-based column index; a fixed band of empty
    trailing columns stays unlocked for free typing / notes.
    """
    from openpyxl.styles import Protection

    unlocked = Protection(locked=False)
    locked = Protection(locked=True)

    used_cols = max(used_cols, 1)
    scan_cols = used_cols + (12 if unlock_columns_from else 0)
    scan_rows = max(ws.max_row or 1, max_row)

    for row in ws.iter_rows(min_row=1, max_row=scan_rows, max_col=used_cols):
        for cell in row:
            cell.protection = locked

    for r, c in unlock_cells or []:
        ws.cell(row=r, column=c).protection = unlocked

    if unlock_columns_from:
        end_col = unlock_columns_from + 11
        for col in range(unlock_columns_from, end_col + 1):
            for row in range(1, scan_rows + 1):
                ws.cell(row=row, column=col).protection = unlocked

    free_start = min(scan_rows + 1, (ws.max_row or 1) + 1)
    for row in range(free_start, free_start + 30):
        for col in range(1, scan_cols + 1):
            ws.cell(row=row, column=col).protection = unlocked

    ws.protection.sheet = True
    ws.protection.enable()
    try:
        ws.protection.set_password(EXCEL_LOCK_PASSWORD)
    except Exception:
        ws.protection.password = EXCEL_LOCK_PASSWORD


def _parse_qty(value) -> float | None:
    from .export_data import parse_money
    raw = str(value or "").strip()
    if not raw:
        return None
    return parse_money(raw)


def _merge_write(ws, row, c1, c2, value, font=None, fill=None, align=None, border=None):
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1 + 1,
                       end_row=row, end_column=c2 + 1)
    cell = ws.cell(row=row, column=c1 + 1, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border
        for c in range(c1, c2 + 1):
            ws.cell(row=row, column=c + 1).border = border
            if fill:
                ws.cell(row=row, column=c + 1).fill = fill
    return cell


def _write_professional_sheet(
    ws, case, form, pairs, rows, *,
    accent=_NAVY,
    subtitle="",
    section_title="",
    pi_totals_block: bool = False,
    price_formulas: bool = False,
    protect: bool = False,
    unlock_unit_price: bool = False,
    hide_client: bool = False,
    hide_project: bool = False,
    blank_unit_price: bool = False,
):
    """Paint brand header, meta boxes, table and footer onto an openpyxl sheet.

    ``pi_totals_block`` — after data rows write Subtotal / VAT / Grand Total (PI).
    ``price_formulas`` — Total Price = Qty × Unit Price; Subtotal = SUM(...).
    ``protect`` — lock sheet with ``EXCEL_LOCK_PASSWORD``.
    ``unlock_unit_price`` — leave Unit Price column editable after protect.
    """
    from openpyxl.utils import get_column_letter
    from .export_data import parse_money, pi_totals, vat_percent

    n_cols = max(len(pairs), 4)
    last_col = n_cols - 1
    mid = max(1, n_cols // 2)
    key_to_col = {key: idx + 1 for idx, (_t, key) in enumerate(pairs)}

    for col_idx, (title, key) in enumerate(pairs):
        width = _COL_WIDTHS.get(key, max(12, min(28, len(title) + 4)))
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = width
    if not pairs:
        for i in range(n_cols):
            ws.column_dimensions[get_column_letter(i + 1)].width = 18

    ws.row_dimensions[1].height = 28
    brand_end = min(2, last_col)
    _merge_write(
        ws, 1, 0, brand_end, "FOOLAD TABAR",
        font=_font(size=18, bold=True, color=accent),
        align=_align("left", "center"),
    )
    kind_label = subtitle or form.get_kind_display().upper()
    kind_start = brand_end + 1
    if kind_start <= last_col:
        _merge_write(
            ws, 1, kind_start, last_col, kind_label,
            font=_font(size=11, bold=True, color=_MUTED),
            align=_align("center", "center"),
        )

    ws.row_dimensions[2].height = 6
    bar = _fill(accent)
    for c in range(n_cols):
        cell = ws.cell(row=2, column=c + 1, value="")
        cell.fill = bar

    doc = _doc_no(case, form)
    form_date = _form_date_jalali(form)
    order = _order_no(case, form) or "—"

    if hide_client:
        # Supply grouped export: no client identity shown anywhere.
        meta = [
            (3, 4, 0, mid - 1, "DOCUMENT NO.", doc),
            (3, 4, mid, last_col, "DATE", form_date),
        ]
        # The grouped export drops the PROJECT box entirely.
        if not hide_project:
            meta.append((5, 6, 0, last_col, "PROJECT", order))
    else:
        client = _client_name_only(case, form)
        meta = [
            (3, 4, 0, mid - 1, "DOCUMENT NO.", doc),
            (3, 4, mid, last_col, "DATE", form_date),
            (5, 6, 0, mid - 1, "CLIENT", client),
        ]
        if not hide_project:
            meta.append((5, 6, mid, last_col, "PROJECT", order))
    box_border = _border(_LINE)
    box_fill = _fill(_SOFT)
    for label_row, value_row, c1, c2, label, value in meta:
        ws.row_dimensions[label_row].height = 14
        ws.row_dimensions[value_row].height = 22
        _merge_write(
            ws, label_row, c1, c2, label,
            font=_font(size=9, bold=True, color=_MUTED),
            align=_align("left", "center"),
        )
        _merge_write(
            ws, value_row, c1, c2, value,
            font=_font(size=11, bold=True, color=_NAVY),
            fill=box_fill, border=box_border,
            align=_align("left", "center"),
        )

    # Header follows the last meta row (so dropping PROJECT doesn't leave a gap).
    meta_last = max((vr for (_lr, vr, _c1, _c2, _l, _v) in meta), default=4)
    spacer = meta_last + 1
    ws.row_dimensions[spacer].height = 8
    header_row = spacer + 1
    if section_title:
        ws.row_dimensions[spacer].height = 20
        _merge_write(
            ws, spacer, 0, last_col, section_title,
            font=_font(size=11, bold=True, color="6B3A8A"),
            fill=_fill("F3E8F8"),
            border=_border(_LINE),
            align=_align("left", "center"),
        )
        ws.row_dimensions[header_row].height = 6
        header_row += 1

    ws.row_dimensions[header_row].height = 22
    hdr_font = _font(size=10, bold=True, color=_WHITE)
    hdr_fill = _fill(accent)
    hdr_align = _align("center", "center")
    hdr_border = _border(_NAVY_DARK if accent == _NAVY else accent)
    for col_idx, (title, _key) in enumerate(pairs):
        cell = ws.cell(row=header_row, column=col_idx + 1, value=title)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = hdr_border

    center_keys = {
        "client_no", "item", "code", "size", "qty", "unit",
        "brand", "time", "unit_price", "service_price", "total_price",
        "unit_svc_price", "total_svc_price",
    }
    wrap_keys = {"desc_client", "desc_ftco", "remark", "comment"}
    data_start = header_row + 1
    grid = _border(_GRID)
    fill_even = _fill(_WHITE)
    fill_odd = _fill(_ZEBRA)
    font_body = _font(size=10, color=_INK)
    unlock_cells: list[tuple[int, int]] = []

    qty_col = key_to_col.get("qty")
    up_col = key_to_col.get("unit_price")
    tp_col = key_to_col.get("total_price")

    for row_idx, row in enumerate(rows or []):
        excel_row = data_start + row_idx
        odd = row_idx % 2 == 1
        bg = fill_odd if odd else fill_even
        if str((row or {}).get("_issue", "") or "") == "1":
            bg = _fill("F3E8F8")
        elif str((row or {}).get("_unsuppliable", "") or "") == "1":
            bg = _fill("FDE8E6")
        elif str((row or {}).get("_service_comment", "") or "").strip().lower() not in (
            "", "nan", "none", "<na>", "null"
        ):
            bg = _fill("EEF0FD")  # SERVICE tag indigo, not Technical Problem purple
        for col_idx, (_title, key) in enumerate(pairs):
            col_no = col_idx + 1
            raw = "" if row is None else str(row.get(key, "") or "")
            flag_label = ""
            flag_color = None
            if key == "desc_ftco" and row and row.get("_flag_label"):
                flag_label = str(row.get("_flag_label") or "")
                if str((row or {}).get("_issue", "") or "") == "1":
                    flag_color = "8E44AD"
                elif str((row or {}).get("_unsuppliable", "") or "") == "1":
                    flag_color = "C0392B"
                elif flag_label == "SERVICE":
                    flag_color = "4456E6"
            value = raw
            number_format = None

            if key == "qty":
                n = _parse_qty(raw)
                value = n if n is not None else raw
            elif key == "unit_price":
                if blank_unit_price:
                    # Grouped supply export: Unit Price is always left empty
                    # (Supply fills it in the workbook).
                    value = ""
                    number_format = "#,##0"
                elif price_formulas:
                    n = parse_money(raw)
                    value = n if raw.strip() else ""
                    if isinstance(value, (int, float)):
                        number_format = "#,##0"
                else:
                    # Keep "1,234 Rial" / "12.50 $" so units match TOTAL PRICE.
                    value = raw
            elif key in ("service_price", "unit_svc_price", "total_svc_price"):
                # Already formatted with currency unit by service_price_export_rows.
                value = raw
            elif key == "total_price":
                if price_formulas and qty_col and up_col:
                    q_letter = get_column_letter(qty_col)
                    u_letter = get_column_letter(up_col)
                    value = f"={q_letter}{excel_row}*{u_letter}{excel_row}"
                    number_format = "#,##0"
                else:
                    value = raw

            if flag_label and isinstance(value, str):
                try:
                    from openpyxl.cell.rich_text import CellRichText, TextBlock
                    from openpyxl.cell.text import InlineFont
                    prefix = (value + "  ") if value.strip() else ""
                    blocks = []
                    if prefix:
                        blocks.append(TextBlock(InlineFont(sz=10, color=_INK), prefix))
                    blocks.append(TextBlock(InlineFont(sz=10, b=True, color=flag_color or _INK),
                                            "[" + flag_label + "]"))
                    value = CellRichText(*blocks)
                except Exception:
                    value = (raw + ("  [" + flag_label + "]" if raw else flag_label)).strip()

            cell = ws.cell(row=excel_row, column=col_no, value=value)
            cell.font = font_body
            cell.fill = bg
            cell.border = grid
            cell.alignment = _align(
                "center" if key in center_keys else "left",
                "center",
                wrap=key in wrap_keys,
            )
            if number_format:
                cell.number_format = number_format
            if unlock_unit_price and key == "unit_price":
                unlock_cells.append((excel_row, col_no))

    n_data = len(rows or [])
    data_end = data_start + n_data - 1 if n_data else data_start - 1
    cursor = data_end + 1

    if pi_totals_block and n_data and tp_col:
        totals = pi_totals(rows)
        label_col = max(1, tp_col - 1)
        value_col = tp_col
        tot_font = _font(size=10, bold=True, color=_NAVY)
        tot_fill = _fill(_SOFT)
        for label, amount in (
            ("SUBTOTAL", totals["subtotal"]),
            *(
                (("TOTAL SERVICE", totals["total_service"]),)
                if totals.get("has_service") else ()
            ),
            (totals.get("vat_label") or f"VAT ({vat_percent():g}%)", totals["vat"]),
            ("GRAND TOTAL", totals["grand_total"]),
        ):
            ws.row_dimensions[cursor].height = 20
            lab = ws.cell(row=cursor, column=label_col, value=label)
            lab.font = tot_font
            lab.fill = tot_fill
            lab.alignment = _align("right", "center")
            lab.border = grid
            num = parse_money(amount)
            val = ws.cell(row=cursor, column=value_col, value=num if amount else 0)
            val.font = tot_font
            val.fill = tot_fill
            val.alignment = _align("center", "center")
            val.border = grid
            val.number_format = "#,##0"
            cursor += 1

    elif price_formulas and n_data and tp_col:
        label_col = max(1, tp_col - 1)
        value_col = tp_col
        tot_font = _font(size=10, bold=True, color=_NAVY)
        tot_fill = _fill(_SOFT)
        ws.row_dimensions[cursor].height = 20
        lab = ws.cell(row=cursor, column=label_col, value="SUBTOTAL")
        lab.font = tot_font
        lab.fill = tot_fill
        lab.alignment = _align("right", "center")
        lab.border = grid
        tp_letter = get_column_letter(tp_col)
        formula = f"=SUM({tp_letter}{data_start}:{tp_letter}{data_end})"
        val = ws.cell(row=cursor, column=value_col, value=formula)
        val.font = tot_font
        val.fill = tot_fill
        val.alignment = _align("center", "center")
        val.border = grid
        val.number_format = "#,##0"
        cursor += 1

    rule_row = cursor
    ws.row_dimensions[rule_row].height = 5
    for c in range(n_cols):
        cell = ws.cell(row=rule_row, column=c + 1, value="")
        cell.fill = _fill(accent)

    footer_row = rule_row + 1
    ws.row_dimensions[footer_row].height = 18
    _merge_write(
        ws, footer_row, 0, max(1, last_col // 2),
        f"Foolad Tabar  ·  {form.get_kind_display()}  ·  {n_data} item(s)",
        font=_font(size=8, color=_FOOTER, italic=True),
        align=_align("left", "center"),
    )
    if last_col >= 2:
        _merge_write(
            ws, footer_row, max(2, last_col - 1), last_col,
            f"Exported {timezone.localtime().strftime('%Y-%m-%d %H:%M')}",
            font=_font(size=8, color=_FOOTER),
            align=_align("right", "center"),
        )

    ws.freeze_panes = "A9"
    ws.print_title_rows = "1:8"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.oddFooter.center.text = "Foolad Tabar Engineering  |  Confidential  |  Page &P of &N"
    ws.oddFooter.center.size = 8

    if protect:
        _protect_worksheet(
            ws,
            unlock_cells=unlock_cells if unlock_unit_price else None,
            unlock_columns_from=(n_cols + 1) if unlock_unit_price else None,
            used_cols=n_cols,
            max_row=max(footer_row + 5, 60),
        )


# ---------------------------------------------------------------------------
# Excel (professional header + table + footer)
# ---------------------------------------------------------------------------
def export_form_excel(case, form, *, hide_client: bool = False, hide_project: bool = False) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from .export_data import build_export_rows, technical_problem_export_rows

    pairs = _export_columns(form)
    rows = build_export_rows(case, form)
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(form)[:31]
    kind = (form.kind or "").upper()
    _write_professional_sheet(
        ws, case, form, pairs, rows,
        pi_totals_block=(kind == FormKind.PI),
        protect=True,
        hide_client=hide_client,
        hide_project=hide_project,
    )

    if kind == FormKind.TO:
        issues = technical_problem_export_rows(form)
        if issues:
            iws = wb.create_sheet(title="Technical Problems")
            issue_pairs = [
                ("Client No.", "client_no"),
                ("Item No.", "item"),
                ("DESCRIPTION CLIENT", "desc_client"),
                ("Technical Problem Detail", "reason"),
            ]
            _write_professional_sheet(
                iws, case, form, issue_pairs, issues,
                subtitle="TECHNICAL PROBLEMS",
                section_title="Technical Problems — flagged rows in this Technical Offer",
                protect=True,
                hide_client=hide_client,
                hide_project=hide_project,
            )

    if kind == FormKind.PI:
        from .export_data import service_price_export_rows
        services = service_price_export_rows(form, case)
        if services:
            sws = wb.create_sheet(title="Services")
            svc_pairs = [
                ("CLIENT ITEM", "client_no"),
                ("FTCO ITEM", "item"),
                ("DESCRIPTION CLIENT", "desc_client"),
                ("SERVICE COMMENT", "comment"),
                ("QTY", "qty"),
                ("UNIT PRICE SERVICE", "unit_svc_price"),
                ("TOTAL PRICE SERVICE", "total_svc_price"),
            ]
            _write_professional_sheet(
                sws, case, form, svc_pairs, services,
                accent="4456E6",
                subtitle="SERVICES",
                section_title="Services — Rows that have services with their prices",
                protect=True,
                hide_client=hide_client,
                hide_project=hide_project,
            )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), export_name_for(case, form) + ".xlsx"


def export_supply_grouped_excel(case, form) -> tuple[bytes, str]:
    """One worksheet per item Group — prices included with Excel formulas.

    Item / Client numbers keep their original values (not restarted per sheet).
    Unit Price is editable; other filled cells are locked. Total Price =
    Qty × Unit Price; Subtotal = SUM of Total Price.
    """
    from openpyxl import Workbook
    from .export_data import build_export_rows

    # Grouped supply export never shows the client's own columns: drop
    # DESCRIPTION CLIENT and REMARK (and brand/time) for every group.
    pairs = [
        (t, k) for t, k in _export_columns(form)
        if k not in {"brand", "time", "desc_client", "remark"}
    ]
    keys = {k for _t, k in pairs}
    if "unit_price" not in keys:
        pairs.append(("UNIT PRICE", "unit_price"))
    if "total_price" not in keys:
        pairs.append(("TOTAL PRICE", "total_price"))

    groups: dict[str, list] = {}
    for row in build_export_rows(case, form):
        key = str(row.get("_group") or "general").strip() or "general"
        groups.setdefault(key, []).append(row)

    wb = Workbook()
    default = wb.active
    first = True
    for group_name, rows in groups.items():
        if first:
            ws = default
            ws.title = group_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=group_name[:31])
        _write_professional_sheet(
            ws, case, form, pairs, rows,
            accent=_TEAL,
            subtitle=f"{form.get_kind_display().upper()}  ·  {group_name}",
            price_formulas=True,
            protect=True,
            unlock_unit_price=True,
            hide_client=True,
            hide_project=True,
            blank_unit_price=True,
        )
    if first:
        _write_professional_sheet(
            default, case, form, pairs, [],
            accent=_TEAL,
            price_formulas=True,
            protect=True,
            unlock_unit_price=True,
            hide_client=True,
            hide_project=True,
            blank_unit_price=True,
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), export_name_for(case, form, group_suffix="groups") + ".xlsx"


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------
def export_form_html(case, form) -> tuple[str, str]:
    from .export_data import build_export_rows

    name = export_name_for(case, form)
    pairs = _export_columns(form)
    columns = [t for t, _k in pairs]
    keys = [_k for _t, _k in pairs]
    rows = build_export_rows(case, form)

    head = f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<title>{html.escape(name)}</title>
<style>
 body{{font-family:'Segoe UI',Calibri,Arial,sans-serif;color:#1B3A4B;margin:24px;}}
 .doc-head{{display:flex;justify-content:space-between;align-items:flex-end;
   border-bottom:3px solid #1B3A4B;padding-bottom:10px;margin-bottom:14px;}}
 .doc-head h1{{font-size:22px;margin:0;letter-spacing:.04em;}}
 .doc-head .kind{{font-size:12px;color:#5A6A7A;font-weight:700;text-transform:uppercase;}}
 .boxes{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:8px 0 16px;}}
 .box{{border:1px solid #D0D7DE;border-radius:8px;padding:8px 12px;background:#F4F7FA;}}
 .box b{{display:block;font-size:10px;color:#5A6A7A;text-transform:uppercase;letter-spacing:.04em;margin-bottom:2px;}}
 table{{width:100%;border-collapse:collapse;font-size:12px;}}
 th{{background:#1B3A4B;color:#fff;padding:7px 6px;border:1px solid #143040;text-align:center;}}
 td{{padding:5px 6px;border:1px solid #D8DEE6;}}
 tr:nth-child(even) td{{background:#F7F9FC;}}
 .foot{{margin-top:18px;padding-top:10px;border-top:1px solid #D0D7DE;
   display:flex;justify-content:space-between;font-size:11px;color:#6B7785;}}
</style></head><body>
<div class="doc-head"><h1>FOOLAD TABAR</h1>
<div class="kind">{html.escape(form.get_kind_display())}</div></div>
<div class="boxes">
 <div class="box"><b>Document No.</b>{html.escape(_doc_no(case, form))}</div>
 <div class="box"><b>Date</b>{html.escape(_form_date_jalali(form))}</div>
 <div class="box"><b>Client</b>{html.escape(_client_name_only(case, form))}</div>
 <div class="box"><b>Project</b>{html.escape(_order_no(case, form) or '—')}</div>
</div>
<table><thead><tr>"""
    head += "".join(f"<th>{html.escape(str(c))}</th>" for c in columns)
    head += "</tr></thead><tbody>"

    body_rows = []
    for row in rows:
        cells = "".join(
            f"<td>{html.escape(str(row.get(k, '') or ''))}</td>" for k in keys
        )
        body_rows.append(f"<tr>{cells}</tr>")
    body = "".join(body_rows) + "</tbody></table>"

    foot = f"""<div class="foot">
<span>Foolad Tabar  ·  {html.escape(form.get_kind_display())}  ·  {len(rows)} item(s)</span>
<span>Exported {html.escape(timezone.localtime().strftime('%Y-%m-%d %H:%M'))}</span>
</div>"""

    return head + body + foot + "</body></html>", name + ".html"


# ---------------------------------------------------------------------------
# PDF — professional print layout (see pdf_export.py)
# ---------------------------------------------------------------------------
def export_form_pdf(case, form, terms: dict | None = None) -> tuple[bytes, str]:
    from .pdf_export import render_form_pdf
    return render_form_pdf(case, form, terms=terms)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _columns_from_table(table) -> list:
    if table and isinstance(table[0], dict):
        return list(table[0].keys())
    return []


def _cell(row: dict, title) -> str:
    value = row.get(title)
    if value is None:
        value = row.get(str(title), "")
    return "" if value is None else str(value)
